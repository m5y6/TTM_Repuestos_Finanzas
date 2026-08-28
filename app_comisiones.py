import os
import warnings
import datetime
import oracledb
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings('ignore', category=UserWarning)

# ==========================================================
# 1. CONFIGURACIÓN DE CONEXIÓN A ORACLE
# ==========================================================
ORACLE_USER = "ttm_admin"
ORACLE_PASS = ""
ORACLE_HOST = "localhost"
ORACLE_PORT = 1521
ORACLE_SERVICE = "xe"

def get_oracle_connection():
    dsn = f"{ORACLE_HOST}:{ORACLE_PORT}/{ORACLE_SERVICE}"
    return oracledb.connect(user=ORACLE_USER, password=ORACLE_PASS, dsn=dsn)

def asegurar_tablas_oracle():
    conn = get_oracle_connection()
    cursor = conn.cursor()
    
    # 1. Tabla de Clientes
    try:
        cursor.execute("""
            CREATE TABLE clientes (
                rut_cliente VARCHAR2(30) PRIMARY KEY,
                nombre      VARCHAR2(150) NOT NULL,
                direccion   VARCHAR2(150),
                comuna      VARCHAR2(50),
                ciudad      VARCHAR2(50),
                giro        VARCHAR2(100),
                telefono    VARCHAR2(50),
                email       VARCHAR2(100)
            )
        """)
        conn.commit()
    except oracledb.DatabaseError:
        pass

    # 2. Columna COMISIONA_INDIVIDUAL en VENDEDORES
    try:
        cursor.execute("""
            ALTER TABLE vendedores ADD (comisiona_individual NUMBER(1) DEFAULT 1 CHECK (comisiona_individual IN (0, 1)))
        """)
        conn.commit()
    except oracledb.DatabaseError:
        pass

    # 3. Tabla de ajustes de comisión
    try:
        cursor.execute("""
            CREATE TABLE comisiones_linea_ajustes (
                tipo_doc     VARCHAR2(5) NOT NULL,
                numero_doc   VARCHAR2(20) NOT NULL,
                cod_producto VARCHAR2(30) NOT NULL,
                cod_vendedor VARCHAR2(10) NOT NULL,
                porc_comis   NUMBER(6,2) NOT NULL,
                CONSTRAINT pk_comis_ajustes PRIMARY KEY (tipo_doc, numero_doc, cod_producto, cod_vendedor)
            )
        """)
        conn.commit()
    except oracledb.DatabaseError:
        pass

    # 4. Tabla de comisión especial
    try:
        cursor.execute("""
            CREATE TABLE comision_especial_global (
                cod_vendedor VARCHAR2(10) PRIMARY KEY,
                nombre       VARCHAR2(100) NOT NULL,
                porc_comis   NUMBER(6,4) DEFAULT 0.25,
                activo       NUMBER(1) DEFAULT 1 CHECK (activo IN (0, 1))
            )
        """)
        conn.commit()
        cursor.execute("""
            INSERT INTO comision_especial_global (cod_vendedor, nombre, porc_comis, activo)
            VALUES ('131', 'FERNANDA DUGAN', 0.25, 1)
        """)
        conn.commit()
    except oracledb.DatabaseError:
        pass

    # 5. Tabla de cuentas TTM Service
    try:
        cursor.execute("""
            CREATE TABLE ttm_service_clientes (
                rut_cliente VARCHAR2(30) PRIMARY KEY,
                nombre      VARCHAR2(150) NOT NULL,
                activo      NUMBER(1) DEFAULT 1 CHECK (activo IN (0, 1))
            )
        """)
        conn.commit()
    except oracledb.DatabaseError:
        pass
    finally:
        cursor.close()
        conn.close()


# ==========================================================
# 2. FUNCIONES GLOBALES DE VALIDACIÓN Y CÁLCULO
# ==========================================================
def normalizar_codigo(val):
    if pd.isna(val) or val is None:
        return 'POZO'
    s = str(val).strip().upper()
    if s.endswith('.0'):
        s = s[:-2]
    return s if s != '' else 'POZO'


def truncar_columnas_df(df, columnas):
    """Trunca los valores monetarios eliminando decimales (sin redondeo hacia arriba)."""
    if df is None or df.empty:
        return df
    df_mod = df.copy()
    for col in columnas:
        if col in df_mod.columns:
            df_mod[col] = df_mod[col].apply(lambda x: int(float(x)) if pd.notna(x) and str(x).strip() != '' else 0)
    return df_mod


def validar_y_normalizar_rango_fechas(f_ini_str, f_fin_str):
    def parse_fecha(txt, nombre_campo):
        valor = txt.strip()
        if not valor:
            return None, f"El campo '{nombre_campo}' no puede estar vacío."
        
        formatos_admitidos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
        for fmt in formatos_admitidos:
            try:
                dt = datetime.datetime.strptime(valor, fmt).date()
                return dt, None
            except ValueError:
                pass

        return None, (
            f"El valor '{valor}' en '{nombre_campo}' no es una fecha válida.\n\n"
            f"Por favor ingrese la fecha completa en formato AAAA-MM-DD (ejemplo: 2026-08-25)."
        )

    dt_ini, err_ini = parse_fecha(f_ini_str, "Fecha Inicio")
    if err_ini:
        return None, None, err_ini

    dt_fin, err_fin = parse_fecha(f_fin_str, "Fecha Fin")
    if err_fin:
        return None, None, err_fin

    if dt_ini > dt_fin:
        return None, None, "La 'Fecha Inicio' no puede ser posterior a la 'Fecha Fin'."

    return dt_ini.strftime("%Y-%m-%d"), dt_fin.strftime("%Y-%m-%d"), None


def extraer_datos_base_oracle(fecha_inicio, fecha_fin):
    conn = get_oracle_connection()

    df_ven_cfg = pd.read_sql_query("SELECT cod_vendedor, nombre, tasa_general, NVL(comisiona_individual, 1) AS comisiona_individual FROM vendedores", conn)
    df_ven_cfg.columns = df_ven_cfg.columns.str.upper()

    df_esp_cfg = pd.read_sql_query("SELECT * FROM comision_especial_global WHERE activo = 1", conn)
    df_esp_cfg.columns = df_esp_cfg.columns.str.upper()

    query = """
    SELECT 
        d.tipo_doc AS tipo_doc_raw,
        CASE 
            WHEN d.tipo_doc = '33' THEN 'Factura'
            WHEN d.tipo_doc = '39' THEN 'Boleta'
            WHEN d.tipo_doc = '61' THEN 'Nota Crédito'
            ELSE d.tipo_doc 
        END AS tipo_doc,
        d.numero_doc AS numero,
        TO_CHAR(d.fecha, 'YYYY-MM-DD') AS fecha_doc,
        CASE 
            WHEN d.tipo_doc = '61' THEN 'ANULADO (NC)'
            WHEN UPPER(TRIM(doc.estado)) = 'CANCELA' THEN 'CANCELADO'
            ELSE 'CRÉDITO (NO PAGADO)'
        END AS estado_doc,
        doc.rut_cliente,
        NVL(c.nombre, NVL(doc.rut_cliente, 'CLIENTE NO REGISTRADO')) AS nombre_cliente,
        d.cod_producto AS producto,
        NVL(p.descripcion, 'SIN DESCRIPCIÓN') AS descripcion_producto,
        (d.cantidad * (CASE WHEN d.tipo_doc = '61' THEN -1.0 ELSE 1.0 END)) AS cant,
        d.precio AS precio_orig,
        NVL(p.precio_costo, 0.0) AS costo_unitario,
        ROUND((NVL(d.descto, 0) + NVL(doc.pdesct1, 0)), 2) AS desc_porc,
        ROUND(d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS precio_final,
        ROUND((d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) * 
        (CASE WHEN d.tipo_doc = '61' THEN -1.0 ELSE 1.0 END), 2) AS tot_venta,
        ROUND((d.cantidad * NVL(p.precio_costo, 0.0)) * 
        (CASE WHEN d.tipo_doc = '61' THEN -1.0 ELSE 1.0 END), 2) AS tot_costo,
        ROUND(((d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) - 
               (d.cantidad * NVL(p.precio_costo, 0.0))) * 
              (CASE WHEN d.tipo_doc = '61' THEN -1.0 ELSE 1.0 END), 2) AS ganancia_total,
        doc.cod_vendedor,
        v.nombre AS nombre_vendedor_original,
        v.tasa_general,
        NVL(v.comisiona_individual, 1) AS comisiona_individual,
        aj.porc_comis AS porc_comis_ajuste,
        CASE 
            WHEN srv.rut_cliente IS NOT NULL THEN 1
            WHEN UPPER(NVL(c.nombre, '')) LIKE '%SERVICE%' OR UPPER(NVL(c.nombre, '')) LIKE '%SERVITECA%' THEN 1
            WHEN UPPER(NVL(doc.rut_cliente, '')) LIKE '%SERVICE%' THEN 1
            ELSE 0
        END AS es_ttm_service
    FROM detalle_venta d
    INNER JOIN documentos_venta doc 
        ON d.tipo_doc = doc.tipo_doc AND d.numero_doc = doc.numero_doc
    LEFT JOIN clientes c 
        ON TRIM(doc.rut_cliente) = TRIM(c.rut_cliente)
        OR REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(doc.rut_cliente)), '.', ''), ' ', ''), '/00', ''), '-', '') = 
           REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(c.rut_cliente)), '.', ''), ' ', ''), '/00', ''), '-', '')
    LEFT JOIN ttm_service_clientes srv
        ON TRIM(doc.rut_cliente) = TRIM(srv.rut_cliente) AND srv.activo = 1
    LEFT JOIN productos p 
        ON d.cod_producto = p.cod_producto
    LEFT JOIN vendedores v 
        ON doc.cod_vendedor = v.cod_vendedor
    LEFT JOIN comisiones_linea_ajustes aj
        ON d.tipo_doc = aj.tipo_doc 
       AND d.numero_doc = aj.numero_doc 
       AND d.cod_producto = aj.cod_producto 
       AND NVL(doc.cod_vendedor, 'POZO') = aj.cod_vendedor
    WHERE d.fecha >= TO_DATE(:f_ini, 'YYYY-MM-DD') 
      AND d.fecha <= TO_DATE(:f_fin, 'YYYY-MM-DD')
    ORDER BY d.fecha DESC, d.numero_doc DESC
    """
    df_raw = pd.read_sql_query(query, conn, params={'f_ini': fecha_inicio, 'f_fin': fecha_fin})
    conn.close()

    if df_raw.empty:
        return pd.DataFrame(), df_ven_cfg, df_esp_cfg

    df_raw.columns = df_raw.columns.str.upper()
    return df_raw, df_ven_cfg, df_esp_cfg


def procesar_calculo_en_memoria(df_raw, df_ven_cfg, df_esp_cfg):
    if df_raw.empty:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame())

    df = df_raw.copy()

    df_ven_cfg_clean = df_ven_cfg.copy()
    df_ven_cfg_clean['COD_VENDEDOR_NORM'] = df_ven_cfg_clean['COD_VENDEDOR'].apply(normalizar_codigo)
    
    dict_ven_ind = dict(zip(df_ven_cfg_clean['COD_VENDEDOR_NORM'], df_ven_cfg_clean['COMISIONA_INDIVIDUAL'].fillna(1).astype(int)))
    dict_ven_nom = dict(zip(df_ven_cfg_clean['COD_VENDEDOR_NORM'], df_ven_cfg_clean['NOMBRE']))
    dict_ven_tasa = dict(zip(df_ven_cfg_clean['COD_VENDEDOR_NORM'], df_ven_cfg_clean['TASA_GENERAL'].fillna(0.02).astype(float)))

    def resolver_vendedor(row):
        if row['ES_TTM_SERVICE'] == 1:
            return 'SERV', 'TTM SERVICE', 0.0

        raw_cod = normalizar_codigo(row['COD_VENDEDOR'])
        comisiona_ind = dict_ven_ind.get(raw_cod, 1)

        if raw_cod == 'POZO' or comisiona_ind == 0:
            cod_final = 'POZO'
            nom_final = 'POZO COMÚN / VENTAS GENERALES'
            tasa_final = 2.0
        else:
            cod_final = raw_cod
            nom_final = dict_ven_nom.get(raw_cod, str(row['NOMBRE_VENDEDOR_ORIGINAL']) if pd.notna(row['NOMBRE_VENDEDOR_ORIGINAL']) else 'POZO COMÚN / VENTAS GENERALES')
            tasa_final = dict_ven_tasa.get(raw_cod, 0.02) * 100.0

        if pd.notna(row['PORC_COMIS_AJUSTE']):
            tasa_final = float(row['PORC_COMIS_AJUSTE'])
        elif str(row['PRODUCTO']).upper().startswith('SER_'):
            tasa_final = 0.0

        return cod_final, nom_final, tasa_final

    res_ven = df.apply(resolver_vendedor, axis=1)
    df['COD_VENDEDOR_RES'] = [r[0] for r in res_ven]
    df['EMPLEADO_RES'] = [r[1] for r in res_ven]
    df['PORC_COMIS_RES'] = [r[2] for r in res_ven]

    df['PAGO_EMPLEADO'] = (df['GANANCIA_TOTAL'] * (df['PORC_COMIS_RES'] / 100.0)).round(2)
    df['GANANCIA_EMPRESA'] = (df['GANANCIA_TOTAL'] - df['PAGO_EMPLEADO']).round(2)
    df['FECHA_CALC'] = datetime.datetime.now().strftime('%Y-%m-%d')

    cols_estandar = [
        'TIPO_DOC', 'NUMERO', 'FECHA_DOC', 'ESTADO_DOC', 'RUT_CLIENTE', 'NOMBRE_CLIENTE',
        'PRODUCTO', 'DESCRIPCION_PRODUCTO', 'CANT', 'PRECIO_ORIG', 'COSTO_UNITARIO', 
        'DESC_PORC', 'PRECIO_FINAL', 'TOT_VENTA', 'TOT_COSTO', 'GANANCIA_TOTAL', 
        'EMPLEADO_RES', 'PORC_COMIS_RES', 'PAGO_EMPLEADO', 'GANANCIA_EMPRESA', 'FECHA_CALC',
        'TIPO_DOC_RAW', 'COD_VENDEDOR_RES', 'ES_TTM_SERVICE'
    ]
    df_total = df[cols_estandar].copy()
    df_total.rename(columns={'EMPLEADO_RES': 'EMPLEADO', 'PORC_COMIS_RES': 'PORC_COMIS', 'COD_VENDEDOR_RES': 'COD_VENDEDOR'}, inplace=True)

    # 1. Comisiones (Vendedores Individuales + Pozo Común)
    df_comis_det = df_total[df_total['ES_TTM_SERVICE'] == 0].copy()

    df_res_ven = df_comis_det.groupby(['COD_VENDEDOR', 'EMPLEADO']).agg(
        TOTAL_VENTAS_NETAS=('TOT_VENTA', lambda x: x[x > 0].sum()),
        TOTAL_NOTAS_CREDITO=('TOT_VENTA', lambda x: abs(x[x < 0].sum())),
        GANANCIA_TOTAL_MARGEN=('GANANCIA_TOTAL', 'sum'),
        TOTAL_COMISION_PAGAR=('PAGO_EMPLEADO', lambda x: max(0, x.sum())),
        GANANCIA_NETA_EMPRESA=('GANANCIA_EMPRESA', 'sum')
    ).reset_index()

    # 2. Comisiones Especiales
    gan_emp_base = df_comis_det['GANANCIA_EMPRESA'].sum()
    lista_esp = []
    if not df_esp_cfg.empty:
        for _, r in df_esp_cfg.iterrows():
            porc = float(r['PORC_COMIS'])
            comis_monto = round(gan_emp_base * (porc / 100.0), 2) if gan_emp_base > 0 else 0.0
            gan_final = round(gan_emp_base - comis_monto, 2)
            lista_esp.append({
                'COD_VENDEDOR': r['COD_VENDEDOR'],
                'BENEFICIARIO': r['NOMBRE'],
                'GANANCIA_BASE_EMPRESA': gan_emp_base,
                'PORC_COMIS_ESPECIAL': porc,
                'COMISION_PAGAR': comis_monto,
                'GANANCIA_FINAL_EMPRESA': gan_final
            })
    df_res_esp = pd.DataFrame(lista_esp)

    # 3. TTM Service
    df_service_det = df_total[df_total['ES_TTM_SERVICE'] == 1].copy()
    if not df_service_det.empty:
        df_service_res = pd.DataFrame([{
            'CONCEPTO': 'TOTAL COMPRAS TTM SERVICE',
            'TOTAL_VENTAS_NETAS': df_service_det[df_service_det['TOT_VENTA'] > 0]['TOT_VENTA'].sum(),
            'TOTAL_NOTAS_CREDITO': abs(df_service_det[df_service_det['TOT_VENTA'] < 0]['TOT_VENTA'].sum()),
            'NETO_FACTURADO': df_service_det['TOT_VENTA'].sum(),
            'COSTO_TOTAL_TRANSFERIDO': df_service_det['TOT_COSTO'].sum(),
            'MARGEN_GANANCIA_TTM': df_service_det['GANANCIA_TOTAL'].sum(),
            'TOTAL_ITEMS': len(df_service_det)
        }])
    else:
        df_service_res = pd.DataFrame()

    return df_res_ven, df_res_esp, df_comis_det, df_service_res, df_service_det


def guardar_ajuste_oracle(tipo_doc_raw, numero_doc, cod_prod, cod_ven, porc_comis):
    conn = get_oracle_connection()
    cursor = conn.cursor()
    merge_sql = """
    MERGE INTO comisiones_linea_ajustes dst
    USING (SELECT :tipo_doc AS tipo_doc, :numero_doc AS numero_doc, :cod_producto AS cod_producto, :cod_vendedor AS cod_vendedor, :porc_comis AS porc_comis FROM dual) src
    ON (dst.tipo_doc = src.tipo_doc AND dst.numero_doc = src.numero_doc AND dst.cod_producto = src.cod_producto AND dst.cod_vendedor = src.cod_vendedor)
    WHEN MATCHED THEN
        UPDATE SET dst.porc_comis = src.porc_comis
    WHEN NOT MATCHED THEN
        INSERT (tipo_doc, numero_doc, cod_producto, cod_vendedor, porc_comis)
        VALUES (src.tipo_doc, src.numero_doc, src.cod_producto, src.cod_vendedor, src.porc_comis)
    """
    cursor.execute(merge_sql, {
        'tipo_doc': tipo_doc_raw,
        'numero_doc': str(numero_doc).strip(),
        'cod_producto': str(cod_prod).strip(),
        'cod_vendedor': str(cod_ven).strip(),
        'porc_comis': float(porc_comis)
    })
    conn.commit()
    cursor.close()
    conn.close()


# ==========================================================
# 3. INTERFAZ GRÁFICA TKINTER (CLASE PRINCIPAL)
# ==========================================================
class AppComisionesOracle(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TTM Repuestos - Finanzas y Control Comercial")
        self.geometry("1340x860")
        self.configure(bg="#f4f6f9")

        asegurar_tablas_oracle()
        
        # Dataframes en memoria
        self.df_raw_cache = pd.DataFrame()
        self.df_ven_cfg_cache = pd.DataFrame()
        self.df_esp_cfg_cache = pd.DataFrame()

        self.df_res_vendedores = None
        self.df_res_especial = None
        self.df_comis_detalle = None
        self.df_service_resumen = None
        self.df_service_detalle = None

        self.columna_orden_actual = "TOTAL_VENTAS_NETAS"
        self.orden_descendente = True

        self.construir_header()
        self.construir_modulos()
        self.verificar_conexion_inicial()

    def obtener_periodo_defecto(self):
        hoy = datetime.date.today()
        if hoy.day >= 25:
            fin = datetime.date(hoy.year, hoy.month, 25)
            if hoy.month == 1:
                inicio = datetime.date(hoy.year - 1, 12, 25)
            else:
                inicio = datetime.date(hoy.year, hoy.month - 1, 25)
        else:
            if hoy.month == 1:
                fin = datetime.date(hoy.year - 1, 12, 25)
                inicio = datetime.date(hoy.year - 1, 11, 25)
            elif hoy.month == 2:
                fin = datetime.date(hoy.year, 1, 25)
                inicio = datetime.date(hoy.year - 1, 12, 25)
            else:
                fin = datetime.date(hoy.year, hoy.month - 1, 25)
                inicio = datetime.date(hoy.year, hoy.month - 2, 25)

        return inicio.strftime("%Y-%m-%d"), fin.strftime("%Y-%m-%d")

    def construir_header(self):
        frame_top = tk.LabelFrame(self, text=" Período Contable y Selección de Temporada (25 al 25) ", font=("Segoe UI", 9, "bold"), bg="#ffffff", padx=15, pady=8)
        frame_top.pack(fill="x", padx=15, pady=8)

        f_ini, f_fin = self.obtener_periodo_defecto()

        tk.Label(frame_top, text="Fecha Inicio:", bg="#ffffff").grid(row=0, column=0, padx=5, sticky="w")
        self.entry_inicio = ttk.Entry(frame_top, width=12)
        self.entry_inicio.insert(0, f_ini)
        self.entry_inicio.grid(row=0, column=1, padx=5)

        tk.Label(frame_top, text="Fecha Fin:", bg="#ffffff").grid(row=0, column=2, padx=10, sticky="w")
        self.entry_fin = ttk.Entry(frame_top, width=12)
        self.entry_fin.insert(0, f_fin)
        self.entry_fin.grid(row=0, column=3, padx=5)

        btn_calc = tk.Button(frame_top, text="⚡ Consultar Oracle", bg="#28a745", fg="white", font=("Segoe UI", 9, "bold"), command=self.ejecutar_consulta_completa)
        btn_calc.grid(row=0, column=4, padx=15)

        btn_exp_gen = tk.Button(frame_top, text="📊 Exportar Excel Comisiones", bg="#17a2b8", fg="white", font=("Segoe UI", 9, "bold"), command=self.exportar_excel_comisiones)
        btn_exp_gen.grid(row=0, column=5, padx=5)

        btn_exp_srv = tk.Button(frame_top, text="🔧 Exportar Excel TTM Service", bg="#6f42c1", fg="white", font=("Segoe UI", 9, "bold"), command=self.exportar_excel_service)
        btn_exp_srv.grid(row=0, column=6, padx=5)

    def construir_modulos(self):
        self.master_notebook = ttk.Notebook(self)
        self.master_notebook.pack(fill="both", expand=True, padx=15, pady=5)

        # MÓDULO 1: COMISIONES
        self.tab_modulo_comisiones = ttk.Frame(self.master_notebook)
        self.master_notebook.add(self.tab_modulo_comisiones, text="  💼 MÓDULO: COMISIONES DE VENDEDORES  ")
        self.construir_subtabs_comisiones()

        # MÓDULO 2: VENTAS TTM SERVICE
        self.tab_modulo_service = ttk.Frame(self.master_notebook)
        self.master_notebook.add(self.tab_modulo_service, text="  🔧 MÓDULO: VENTAS TTM A TTM SERVICE  ")
        self.construir_subtabs_service()

        self.lbl_estado = tk.Label(self, text="Conectando a Oracle...", bd=1, relief="sunken", anchor="w", bg="#e9ecef", font=("Segoe UI", 9))
        self.lbl_estado.pack(side="bottom", fill="x")

    def construir_subtabs_comisiones(self):
        self.nb_comis = ttk.Notebook(self.tab_modulo_comisiones)
        self.nb_comis.pack(fill="both", expand=True, padx=5, pady=5)

        self.subtab_resumen_c = ttk.Frame(self.nb_comis)
        self.nb_comis.add(self.subtab_resumen_c, text=" 📊 Resumen Liquidación y Empresa ")
        self.crear_vista_resumen_comisiones()

        self.subtab_detalle_c = ttk.Frame(self.nb_comis)
        self.nb_comis.add(self.subtab_detalle_c, text=" 🔍 Detalle Transacciones (Historial Completo) ")
        self.crear_vista_detalle_comisiones()

        self.subtab_config_c = ttk.Frame(self.nb_comis)
        self.nb_comis.add(self.subtab_config_c, text=" ⚙️ Panel de Configuración General ")
        self.crear_vista_config_especial()

    def crear_vista_resumen_comisiones(self):
        frame_sort_bar = tk.Frame(self.subtab_resumen_c, bg="#ffffff", bd=1, relief="groove", padx=10, pady=5)
        frame_sort_bar.pack(fill="x", padx=10, pady=(6, 2))

        tk.Label(frame_sort_bar, text="Ordenar por:", font=("Segoe UI", 9, "bold"), bg="#ffffff").pack(side="left", padx=5)

        btn_sort_v = ttk.Button(frame_sort_bar, text="🔝 Mayor Venta", command=lambda: self.ordenar_resumen_por("TOTAL_VENTAS_NETAS"))
        btn_sort_v.pack(side="left", padx=3)

        btn_sort_nc = ttk.Button(frame_sort_bar, text="⚠️ Mayor Nota Crédito", command=lambda: self.ordenar_resumen_por("TOTAL_NOTAS_CREDITO"))
        btn_sort_nc.pack(side="left", padx=3)

        btn_sort_m = ttk.Button(frame_sort_bar, text="💰 Mayor Margen", command=lambda: self.ordenar_resumen_por("GANANCIA_TOTAL_MARGEN"))
        btn_sort_m.pack(side="left", padx=3)

        btn_sort_c = ttk.Button(frame_sort_bar, text="💵 Mayor Comisión", command=lambda: self.ordenar_resumen_por("TOTAL_COMISION_PAGAR"))
        btn_sort_c.pack(side="left", padx=3)

        btn_sort_e = ttk.Button(frame_sort_bar, text="🏢 Mayor Ganancia Empresa", command=lambda: self.ordenar_resumen_por("GANANCIA_NETA_EMPRESA"))
        btn_sort_e.pack(side="left", padx=3)

        self.lbl_info_orden = tk.Label(frame_sort_bar, text="Ordenado: Mayor Venta Neta ▼", font=("Segoe UI", 8, "italic"), bg="#ffffff", fg="#6c757d")
        self.lbl_info_orden.pack(side="right", padx=10)

        frame_v = ttk.Frame(self.subtab_resumen_c)
        frame_v.pack(fill="both", expand=True, padx=10, pady=2)

        cols_v = ("COD", "EMPLEADO", "VENTAS_NETAS", "NOTAS_CREDITO", "GANANCIA_TOTAL", "PAGO_EMPLEADO", "GANANCIA_EMPRESA")
        self.tree_res_v = ttk.Treeview(frame_v, columns=cols_v, show="headings", height=7)
        
        self.tree_res_v.heading("COD", text="Cód", command=lambda: self.ordenar_resumen_por("COD_VENDEDOR"))
        self.tree_res_v.heading("EMPLEADO", text="Vendedor / Entidad", command=lambda: self.ordenar_resumen_por("EMPLEADO"))
        self.tree_res_v.heading("VENTAS_NETAS", text="Ventas Netas ($)", command=lambda: self.ordenar_resumen_por("TOTAL_VENTAS_NETAS"))
        self.tree_res_v.heading("NOTAS_CREDITO", text="Notas Crédito ($)", command=lambda: self.ordenar_resumen_por("TOTAL_NOTAS_CREDITO"))
        self.tree_res_v.heading("GANANCIA_TOTAL", text="Ganancia / Margen ($)", command=lambda: self.ordenar_resumen_por("GANANCIA_TOTAL_MARGEN"))
        self.tree_res_v.heading("PAGO_EMPLEADO", text="Pago Comisión ($)", command=lambda: self.ordenar_resumen_por("TOTAL_COMISION_PAGAR"))
        self.tree_res_v.heading("GANANCIA_EMPRESA", text="Ganancia Empresa ($)", command=lambda: self.ordenar_resumen_por("GANANCIA_NETA_EMPRESA"))

        self.tree_res_v.column("COD", width=60, anchor="center")
        self.tree_res_v.column("EMPLEADO", width=250, anchor="w")
        self.tree_res_v.column("VENTAS_NETAS", width=140, anchor="e")
        self.tree_res_v.column("NOTAS_CREDITO", width=140, anchor="e")
        self.tree_res_v.column("GANANCIA_TOTAL", width=150, anchor="e")
        self.tree_res_v.column("PAGO_EMPLEADO", width=150, anchor="e")
        self.tree_res_v.column("GANANCIA_EMPRESA", width=150, anchor="e")

        self.tree_res_v.tag_configure('fila_pozo', background='#E2E3E5', font=("Segoe UI", 9, "bold"))
        self.tree_res_v.tag_configure('fila_normal', background='#FFFFFF')

        sb_v = ttk.Scrollbar(frame_v, orient="vertical", command=self.tree_res_v.yview)
        self.tree_res_v.configure(yscrollcommand=sb_v.set)
        self.tree_res_v.pack(side="left", fill="both", expand=True)
        sb_v.pack(side="right", fill="y")

        # PANEL CONSOLIDADO DINÁMICO
        self.frame_totales_contador = tk.Frame(self.subtab_resumen_c, bg="#212529", padx=15, pady=10)
        self.frame_totales_contador.pack(fill="x", padx=10, pady=(2, 8))

        self.lbl_tot_contador = tk.Label(
            self.frame_totales_contador, 
            text="TOTAL A LIQUIDAR: $0  |  MARGEN TOTAL: $0  |  GANANCIA NETA EMPRESA: $0", 
            font=("Segoe UI", 9, "bold"), 
            bg="#212529", 
            fg="#28A745"
        )
        self.lbl_tot_contador.pack(anchor="w")

        # Tabla 2: Comisión Especial
        lbl_e = tk.Label(self.subtab_resumen_c, text="2. COMISIÓN ESPECIAL (CALCULADA SOBRE LA GANANCIA LÍQUIDA DE LA EMPRESA)", font=("Segoe UI", 9, "bold"), fg="#0056b3", anchor="w")
        lbl_e.pack(fill="x", padx=10, pady=(4, 2))

        frame_e = ttk.Frame(self.subtab_resumen_c)
        frame_e.pack(fill="x", padx=10, pady=2)

        cols_e = ("COD", "BENEFICIARIO", "GANANCIA_BASE", "PORC_APLICADO", "COMISION_PAGAR", "GANANCIA_FINAL")
        self.tree_res_e = ttk.Treeview(frame_e, columns=cols_e, show="headings", height=3)

        self.tree_res_e.heading("COD", text="Cód")
        self.tree_res_e.heading("BENEFICIARIO", text="Beneficiario Comisión Especial")
        self.tree_res_e.heading("GANANCIA_BASE", text="Ganancia Base Empresa ($)")
        self.tree_res_e.heading("PORC_APLICADO", text="% Asignado")
        self.tree_res_e.heading("COMISION_PAGAR", text="Comisión a Pagar ($)")
        self.tree_res_e.heading("GANANCIA_FINAL", text="Ganancia Neta Final Empresa ($)")

        self.tree_res_e.column("COD", width=60, anchor="center")
        self.tree_res_e.column("BENEFICIARIO", width=250, anchor="w")
        self.tree_res_e.column("GANANCIA_BASE", width=180, anchor="e")
        self.tree_res_e.column("PORC_APLICADO", width=110, anchor="center")
        self.tree_res_e.column("COMISION_PAGAR", width=160, anchor="e")
        self.tree_res_e.column("GANANCIA_FINAL", width=200, anchor="e")

        self.tree_res_e.pack(fill="x")

    def ordenar_resumen_por(self, col):
        if self.df_res_vendedores is None or self.df_res_vendedores.empty:
            return

        if self.columna_orden_actual == col:
            self.orden_descendente = not self.orden_descendente
        else:
            self.columna_orden_actual = col
            self.orden_descendente = True

        simbolo = "▼ (Mayor a Menor)" if self.orden_descendente else "▲ (Menor a Mayor)"
        self.lbl_info_orden.config(text=f"Ordenado por: {col} {simbolo}")
        self.repoblar_tabla_resumen()

    def repoblar_tabla_resumen(self):
        if self.df_res_vendedores is None or self.df_res_vendedores.empty:
            return

        df_sorted = self.df_res_vendedores.sort_values(by=self.columna_orden_actual, ascending=not self.orden_descendente)

        for i in self.tree_res_v.get_children():
            self.tree_res_v.delete(i)

        for _, r in df_sorted.iterrows():
            tag = 'fila_pozo' if r['COD_VENDEDOR'] == 'POZO' else 'fila_normal'
            self.tree_res_v.insert("", "end", values=(
                r['COD_VENDEDOR'],
                r['EMPLEADO'],
                f"${r['TOTAL_VENTAS_NETAS']:,.0f}".replace(",", "."),
                f"${r['TOTAL_NOTAS_CREDITO']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_TOTAL_MARGEN']:,.0f}".replace(",", "."),
                f"${r['TOTAL_COMISION_PAGAR']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_NETA_EMPRESA']:,.0f}".replace(",", ".")
            ), tags=(tag,))

    def crear_vista_detalle_comisiones(self):
        frame_filtros = tk.Frame(self.subtab_detalle_c, bg="#ffffff", bd=1, relief="groove", padx=10, pady=6)
        frame_filtros.pack(fill="x", padx=5, pady=5)

        tk.Label(frame_filtros, text="🔍 Buscar N°:", bg="#ffffff", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(5, 2))
        self.entry_busq_c = ttk.Entry(frame_filtros, width=15)
        self.entry_busq_c.pack(side="left", padx=(0, 15))
        self.entry_busq_c.bind("<KeyRelease>", lambda event: self.aplicar_filtros_comisiones())

        tk.Label(frame_filtros, text="👤 Empleado:", bg="#ffffff", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(5, 2))
        self.combo_emp_c = ttk.Combobox(frame_filtros, width=28, state="readonly")
        self.combo_emp_c.set("TODOS")
        self.combo_emp_c.pack(side="left", padx=(0, 15))
        self.combo_emp_c.bind("<<ComboboxSelected>>", lambda event: self.aplicar_filtros_comisiones())

        btn_limp = ttk.Button(frame_filtros, text="🧹 Limpiar", command=self.limpiar_filtros_comisiones)
        btn_limp.pack(side="left", padx=5)

        btn_edit_c = tk.Button(frame_filtros, text="✏️ Modificar % Ganancia Fila", bg="#ffc107", fg="black", font=("Segoe UI", 8, "bold"), command=self.modal_editar_comision_seleccion)
        btn_edit_c.pack(side="left", padx=15)

        lbl_leyenda = tk.Label(frame_filtros, text="  Amarillo = NOTA DE CRÉDITO (Doble clic para ver Evidencia de Origen)  ", bg="#FFF3CD", fg="#856404", font=("Segoe UI", 8, "bold"), bd=1, relief="solid")
        lbl_leyenda.pack(side="left", padx=15)

        self.lbl_conteo_c = tk.Label(frame_filtros, text="", bg="#ffffff", fg="#6c757d", font=("Segoe UI", 9, "italic"))
        self.lbl_conteo_c.pack(side="right", padx=10)

        frame_grid = ttk.Frame(self.subtab_detalle_c)
        frame_grid.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # NUEVO ORDEN: Se eliminó Descripción y Vendedor está al lado de Fecha Doc
        self.cols_det_c = [
            ("TIPO_DOC", "Tipo Doc", 90, "center"),
            ("NUMERO", "Número", 85, "center"),
            ("FECHA_DOC", "Fecha Doc", 95, "center"),
            ("EMPLEADO", "Vendedor", 160, "w"),
            ("ESTADO_DOC", "Estado Doc", 155, "center"),
            ("NOMBRE_CLIENTE", "Cliente Comprador", 180, "w"),
            ("PRODUCTO", "Producto", 110, "w"),
            ("CANT", "Cant", 55, "center"),
            ("PRECIO_ORIG", "Precio Orig ($)", 105, "e"),
            ("COSTO_UNITARIO", "Costo ($)", 105, "e"),
            ("DESC_PORC", "Desc %", 70, "center"),
            ("PRECIO_FINAL", "Precio Final ($)", 110, "e"),
            ("TOT_VENTA", "Tot Venta ($)", 115, "e"),
            ("TOT_COSTO", "Tot Costo ($)", 115, "e"),
            ("GANANCIA_TOTAL", "Ganancia Total ($)", 125, "e"),
            ("PORC_COMIS", "% Comis", 80, "center"),
            ("PAGO_EMPLEADO", "Pago Empleado ($)", 125, "e"),
            ("GANANCIA_EMPRESA", "Ganancia Empresa ($)", 135, "e")
        ]

        col_ids = [c[0] for c in self.cols_det_c]
        self.tree_det_c = ttk.Treeview(frame_grid, columns=col_ids, show="headings", selectmode="extended")

        for col_id, titulo, ancho, alineacion in self.cols_det_c:
            self.tree_det_c.heading(col_id, text=titulo)
            self.tree_det_c.column(col_id, width=ancho, anchor=alineacion)

        self.tree_det_c.tag_configure('nota_credito', background='#FFF3CD', foreground='#856404')
        self.tree_det_c.tag_configure('fila_par', background='#F8F9FA')
        self.tree_det_c.tag_configure('fila_impar', background='#FFFFFF')

        sb_y = ttk.Scrollbar(frame_grid, orient="vertical", command=self.tree_det_c.yview)
        sb_x = ttk.Scrollbar(frame_grid, orient="horizontal", command=self.tree_det_c.xview)
        
        self.tree_det_c.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        self.tree_det_c.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")

        frame_grid.grid_rowconfigure(0, weight=1)
        frame_grid.grid_columnconfigure(0, weight=1)

        self.tree_det_c.bind("<Double-1>", self.on_double_click_detalle_comisiones)

    def on_double_click_detalle_comisiones(self, event):
        seleccionados = self.tree_det_c.selection()
        if not seleccionados:
            return

        item_id = seleccionados[0]
        vals = self.tree_det_c.item(item_id, "values")
        tipo_doc, num_doc, fecha, emp, estado, cli, prod = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6]

        match = self.df_comis_detalle[(self.df_comis_detalle['NUMERO'] == num_doc) & (self.df_comis_detalle['PRODUCTO'] == prod)]
        if match.empty:
            return
        fila = match.iloc[0]

        if fila['TIPO_DOC_RAW'] == '61' or tipo_doc == 'Nota Crédito':
            self.modal_auditoria_nota_credito(fila)
        else:
            self.modal_visor_factura_con_nc(fila)

    def modal_auditoria_nota_credito(self, fila):
        num_nc = str(fila['NUMERO']).strip()
        rut_cli = str(fila['RUT_CLIENTE']).strip() if pd.notna(fila['RUT_CLIENTE']) else ""
        nom_cli = fila['NOMBRE_CLIENTE']
        fecha_nc = fila['FECHA_DOC']
        vendedor_nc = fila['EMPLEADO']

        win = tk.Toplevel(self)
        win.title(f"Auditoría y Evidencia: Nota de Crédito N° {num_nc}")
        win.geometry("1020x680")
        win.transient(self)
        win.grab_set()

        frame_head = tk.Frame(win, bg="#FFF3CD", bd=1, relief="solid", padx=15, pady=10)
        frame_head.pack(fill="x", padx=10, pady=(10, 5))

        tk.Label(frame_head, text=f"🔍 AUDITORÍA DE ANULACIÓN - NOTA DE CRÉDITO N° {num_nc}", font=("Segoe UI", 11, "bold"), bg="#FFF3CD", fg="#856404").pack(anchor="w")
        tk.Label(frame_head, text=f"Fecha de Emisión: {fecha_nc}   |   Vendedor Emisor: {vendedor_nc}   |   Estado: ANULADO / DEVOLUCIÓN", font=("Segoe UI", 9), bg="#FFF3CD").pack(anchor="w", pady=(2, 0))
        tk.Label(frame_head, text=f"Cliente: {nom_cli} (RUT: {rut_cli})", font=("Segoe UI", 9, "bold"), bg="#FFF3CD").pack(anchor="w", pady=(2, 0))

        lbl_s1 = tk.Label(win, text="1. ÍTEMS DEVUELTOS / RESTADOS EN ESTA NOTA DE CRÉDITO:", font=("Segoe UI", 9, "bold"), fg="#333", anchor="w")
        lbl_s1.pack(fill="x", padx=10, pady=(5, 2))

        cols_nc = ("PRODUCTO", "CANT_DEV", "PRECIO", "COSTO", "NETO_ANULADO", "MARGEN_RESTADO")
        tree_nc_items = ttk.Treeview(win, columns=cols_nc, show="headings", height=4)
        tree_nc_items.heading("PRODUCTO", text="Cód Producto")
        tree_nc_items.heading("CANT_DEV", text="Cant Devuelta")
        tree_nc_items.heading("PRECIO", text="Precio Unit ($)")
        tree_nc_items.heading("COSTO", text="Costo Unit ($)")
        tree_nc_items.heading("NETO_ANULADO", text="Venta Restada ($)")
        tree_nc_items.heading("MARGEN_RESTADO", text="Margen Restado ($)")

        tree_nc_items.column("PRODUCTO", width=140, anchor="w")
        tree_nc_items.column("CANT_DEV", width=100, anchor="center")
        tree_nc_items.column("PRECIO", width=120, anchor="e")
        tree_nc_items.column("COSTO", width=120, anchor="e")
        tree_nc_items.column("NETO_ANULADO", width=130, anchor="e")
        tree_nc_items.column("MARGEN_RESTADO", width=130, anchor="e")

        try:
            conn = get_oracle_connection()
            query_nc = """
            SELECT 
                d.cod_producto,
                ABS(d.cantidad) AS cant_dev,
                d.precio,
                NVL(p.precio_costo, 0.0) AS costo,
                ROUND(ABS(d.cantidad) * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS neto_anulado,
                ROUND((ABS(d.cantidad) * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) - 
                      (ABS(d.cantidad) * NVL(p.precio_costo, 0.0)), 2) AS margen_restado
            FROM detalle_venta d
            INNER JOIN documentos_venta doc 
                ON d.tipo_doc = doc.tipo_doc AND d.numero_doc = doc.numero_doc
            LEFT JOIN productos p 
                ON d.cod_producto = p.cod_producto
            WHERE d.tipo_doc = '61' 
              AND (TRIM(d.numero_doc) = :num_raw OR LTRIM(TRIM(d.numero_doc), '0') = :num_clean)
            ORDER BY d.cod_producto
            """
            df_nc_db = pd.read_sql_query(query_nc, conn, params={
                'num_raw': num_nc,
                'num_clean': num_nc.lstrip('0')
            })
            conn.close()
            df_nc_db.columns = df_nc_db.columns.str.upper()

            for _, r in df_nc_db.iterrows():
                tree_nc_items.insert("", "end", values=(
                    r['COD_PRODUCTO'],
                    f"{r['CANT_DEV']:.0f}",
                    f"${r['PRECIO']:,.0f}".replace(",", "."),
                    f"${r['COSTO']:,.0f}".replace(",", "."),
                    f"${r['NETO_ANULADO']:,.0f}".replace(",", "."),
                    f"${r['MARGEN_RESTADO']:,.0f}".replace(",", ".")
                ))
        except Exception as e:
            tree_nc_items.insert("", "end", values=("ERROR", "", "", "", "", ""))

        tree_nc_items.pack(fill="x", padx=10, pady=2)

        nb_evidencia = ttk.Notebook(win)
        nb_evidencia.pack(fill="both", expand=True, padx=10, pady=(8, 10))

        tab_origen = ttk.Frame(nb_evidencia)
        nb_evidencia.add(tab_origen, text="  📄 Factura / Boleta Específica de Origen  ")

        tab_historial_cli = ttk.Frame(nb_evidencia)
        nb_evidencia.add(tab_historial_cli, text="  📚 Historial Completo de Compras del Cliente  ")

        # TAB A: DOCUMENTO DE ORIGEN ESPECÍFICO
        frame_busq_orig = tk.Frame(tab_origen, padx=10, pady=6)
        frame_busq_orig.pack(fill="x")

        tk.Label(frame_busq_orig, text="N° Factura / Boleta de Origen:", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        entry_doc_especifico = ttk.Entry(frame_busq_orig, width=15)
        entry_doc_especifico.pack(side="left", padx=5)

        lbl_estado_origen = tk.Label(frame_busq_orig, text="", font=("Segoe UI", 9, "italic"), fg="#0056b3")
        lbl_estado_origen.pack(side="right", padx=10)

        cols_doc_especifico = ("TIPO", "NUMERO", "FECHA", "VENDEDOR", "PRODUCTO", "CANT", "PRECIO_FINAL", "TOTAL_VENTA", "MARGEN")
        tree_doc_especifico = ttk.Treeview(tab_origen, columns=cols_doc_especifico, show="headings", height=6)
        
        tree_doc_especifico.heading("TIPO", text="Tipo")
        tree_doc_especifico.heading("NUMERO", text="N° Docto")
        tree_doc_especifico.heading("FECHA", text="Fecha")
        tree_doc_especifico.heading("VENDEDOR", text="Vendedor Original")
        tree_doc_especifico.heading("PRODUCTO", text="Producto")
        tree_doc_especifico.heading("CANT", text="Cant Vendida")
        tree_doc_especifico.heading("PRECIO_FINAL", text="Precio Cobrado ($)")
        tree_doc_especifico.heading("TOTAL_VENTA", text="Total Venta ($)")
        tree_doc_especifico.heading("MARGEN", text="Margen ($)")

        tree_doc_especifico.column("TIPO", width=75, anchor="center")
        tree_doc_especifico.column("NUMERO", width=80, anchor="center")
        tree_doc_especifico.column("FECHA", width=85, anchor="center")
        tree_doc_especifico.column("VENDEDOR", width=140, anchor="w")
        tree_doc_especifico.column("PRODUCTO", width=110, anchor="w")
        tree_doc_especifico.column("CANT", width=85, anchor="center")
        tree_doc_especifico.column("PRECIO_FINAL", width=110, anchor="e")
        tree_doc_especifico.column("TOTAL_VENTA", width=115, anchor="e")
        tree_doc_especifico.column("MARGEN", width=115, anchor="e")

        def cargar_documento_origen_especifico(num_buscar):
            num_clean = num_buscar.strip().lstrip('0')
            if not num_clean:
                messagebox.showwarning("Atención", "Ingrese el número de la Factura o Boleta de origen.", parent=win)
                return

            for item in tree_doc_especifico.get_children():
                tree_doc_especifico.delete(item)

            try:
                conn = get_oracle_connection()
                q_orig = """
                SELECT 
                    CASE WHEN d.tipo_doc = '33' THEN 'Factura' WHEN d.tipo_doc = '39' THEN 'Boleta' ELSE d.tipo_doc END AS tipo_doc,
                    d.numero_doc AS numero,
                    TO_CHAR(d.fecha, 'YYYY-MM-DD') AS fecha_doc,
                    NVL(v.nombre, 'SIN VENDEDOR') AS vendedor_doc,
                    d.cod_producto,
                    d.cantidad,
                    ROUND(d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS precio_final,
                    ROUND(d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS tot_venta,
                    ROUND((d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) - (d.cantidad * NVL(p.precio_costo, 0.0)), 2) AS ganancia
                FROM detalle_venta d
                INNER JOIN documentos_venta doc 
                    ON d.tipo_doc = doc.tipo_doc AND d.numero_doc = doc.numero_doc
                LEFT JOIN vendedores v ON doc.cod_vendedor = v.cod_vendedor
                LEFT JOIN productos p ON d.cod_producto = p.cod_producto
                WHERE d.tipo_doc IN ('33', '39')
                  AND (TRIM(d.numero_doc) = :num_raw OR LTRIM(TRIM(d.numero_doc), '0') = :num_clean)
                ORDER BY d.cod_producto
                """
                df_orig = pd.read_sql_query(q_orig, conn, params={
                    'num_raw': num_buscar.strip(),
                    'num_clean': num_clean
                })
                conn.close()
                df_orig.columns = df_orig.columns.str.upper()

                if df_orig.empty:
                    tree_doc_especifico.insert("", "end", values=("INFO", "-", "-", "-", f"No se encontró la Factura/Boleta N° {num_buscar}", "", "", "", ""))
                    lbl_estado_origen.config(text=f"No encontrada: {num_buscar}")
                else:
                    tot_v = df_orig['TOT_VENTA'].sum()
                    tot_m = df_orig['GANANCIA'].sum()
                    lbl_estado_origen.config(text=f"Venta Total Original: ${tot_v:,.0f} | Margen Original: ${tot_m:,.0f}".replace(",", "."))
                    
                    for _, r in df_orig.iterrows():
                        tree_doc_especifico.insert("", "end", values=(
                            r['TIPO_DOC'],
                            r['NUMERO'],
                            r['FECHA_DOC'],
                            r['VENDEDOR_DOC'],
                            r['COD_PRODUCTO'],
                            f"{r['CANTIDAD']:.0f}",
                            f"${r['PRECIO_FINAL']:,.0f}".replace(",", "."),
                            f"${r['TOT_VENTA']:,.0f}".replace(",", "."),
                            f"${r['GANANCIA']:,.0f}".replace(",", ".")
                        ))
                    nb_evidencia.select(tab_origen)
            except Exception as e:
                tree_doc_especifico.insert("", "end", values=("ERROR", "-", "-", "-", f"Error: {e}", "", "", "", ""))

        btn_ver_origen = ttk.Button(frame_busq_orig, text="🔍 Cargar Documento de Origen", command=lambda: cargar_documento_origen_especifico(entry_doc_especifico.get()))
        btn_ver_origen.pack(side="left", padx=5)

        tree_doc_especifico.pack(fill="both", expand=True, padx=5, pady=5)

        # TAB B: HISTORIAL COMPLETO DE COMPRAS DEL CLIENTE
        frame_hist_top = tk.Frame(tab_historial_cli, padx=10, pady=6)
        frame_hist_top.pack(fill="x")

        tk.Label(frame_hist_top, text="Todas las Facturas y Boletas emitidas a este cliente (Doble clic para cargar como origen):", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        
        cols_hist = ("TIPO", "NUMERO", "FECHA", "PRODUCTO", "CANT", "PRECIO_FINAL", "TOTAL_VENTA", "MARGEN")
        tree_hist_cli = ttk.Treeview(tab_historial_cli, columns=cols_hist, show="headings", height=6)
        
        tree_hist_cli.heading("TIPO", text="Tipo")
        tree_hist_cli.heading("NUMERO", text="N° Docto")
        tree_hist_cli.heading("FECHA", text="Fecha")
        tree_hist_cli.heading("PRODUCTO", text="Producto")
        tree_hist_cli.heading("CANT", text="Cant")
        tree_hist_cli.heading("PRECIO_FINAL", text="Precio Unit ($)")
        tree_hist_cli.heading("TOTAL_VENTA", text="Total Venta ($)")
        tree_hist_cli.heading("MARGEN", text="Margen ($)")

        tree_hist_cli.column("TIPO", width=80, anchor="center")
        tree_hist_cli.column("NUMERO", width=85, anchor="center")
        tree_hist_cli.column("FECHA", width=90, anchor="center")
        tree_hist_cli.column("PRODUCTO", width=110, anchor="w")
        tree_hist_cli.column("CANT", width=55, anchor="center")
        tree_hist_cli.column("PRECIO_FINAL", width=105, anchor="e")
        tree_hist_cli.column("TOTAL_VENTA", width=110, anchor="e")
        tree_hist_cli.column("MARGEN", width=110, anchor="e")

        def seleccionar_desde_historial(event):
            sel = tree_hist_cli.selection()
            if not sel:
                return
            vals = tree_hist_cli.item(sel[0], "values")
            num_sel = vals[1]
            if num_sel and num_sel != "-":
                entry_doc_especifico.delete(0, tk.END)
                entry_doc_especifico.insert(0, num_sel)
                cargar_documento_origen_especifico(num_sel)

        tree_hist_cli.bind("<Double-1>", seleccionar_desde_historial)

        btn_cargar_hist = ttk.Button(frame_hist_top, text="👁️ Cargar Factura Seleccionada", command=lambda: seleccionar_desde_historial(None))
        btn_cargar_hist.pack(side="right", padx=5)

        try:
            conn = get_oracle_connection()
            clean_rut = str(rut_cli).upper().strip().replace('.', '').replace(' ', '').replace('/00', '').replace('-', '')
            q_hist = """
            SELECT 
                CASE WHEN d.tipo_doc = '33' THEN 'Factura' WHEN d.tipo_doc = '39' THEN 'Boleta' ELSE d.tipo_doc END AS tipo_doc,
                d.numero_doc AS numero,
                TO_CHAR(d.fecha, 'YYYY-MM-DD') AS fecha_doc,
                d.cod_producto,
                d.cantidad,
                ROUND(d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS precio_final,
                ROUND(d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS tot_venta,
                ROUND((d.cantidad * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) - (d.cantidad * NVL(p.precio_costo, 0.0)), 2) AS ganancia
            FROM detalle_venta d
            INNER JOIN documentos_venta doc 
                ON d.tipo_doc = doc.tipo_doc AND d.numero_doc = doc.numero_doc
            LEFT JOIN productos p ON d.cod_producto = p.cod_producto
            WHERE d.tipo_doc IN ('33', '39')
              AND (
                  TRIM(doc.rut_cliente) = :rut_raw 
                  OR REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(doc.rut_cliente)), '.', ''), ' ', ''), '/00', ''), '-', '') = :rut_clean
              )
            ORDER BY d.fecha DESC, d.numero_doc DESC
            """
            df_hist = pd.read_sql_query(q_hist, conn, params={
                'rut_raw': str(rut_cli).strip(),
                'rut_clean': clean_rut
            })
            conn.close()
            df_hist.columns = df_hist.columns.str.upper()

            if df_hist.empty:
                tree_hist_cli.insert("", "end", values=("INFO", "-", "-", "Sin compras previas registradas", "", "", "", ""))
            else:
                for _, r in df_hist.iterrows():
                    tree_hist_cli.insert("", "end", values=(
                        r['TIPO_DOC'],
                        r['NUMERO'],
                        r['FECHA_DOC'],
                        r['COD_PRODUCTO'],
                        f"{r['CANTIDAD']:.0f}",
                        f"${r['PRECIO_FINAL']:,.0f}".replace(",", "."),
                        f"${r['TOT_VENTA']:,.0f}".replace(",", "."),
                        f"${r['GANANCIA']:,.0f}".replace(",", ".")
                    ))
                
                primera_factura = str(df_hist.iloc[0]['NUMERO'])
                entry_doc_especifico.insert(0, primera_factura)
                cargar_documento_origen_especifico(primera_factura)

        except Exception as e:
            tree_hist_cli.insert("", "end", values=("ERROR", "-", "-", f"Error: {e}", "", "", "", ""))

        tree_hist_cli.pack(fill="both", expand=True, padx=5, pady=5)

    # ==========================================================
    # MODAL: VISOR DE FACTURA / BOLETA CON CRUCE DE NOTAS DE CRÉDITO
    # ==========================================================
    def modal_visor_factura_con_nc(self, fila):
        tipo_doc = fila['TIPO_DOC']
        num_doc = str(fila['NUMERO']).strip()
        rut_cli = str(fila['RUT_CLIENTE']).strip() if pd.notna(fila['RUT_CLIENTE']) else ""
        nom_cli = fila['NOMBRE_CLIENTE']
        fecha_doc = fila['FECHA_DOC']
        vendedor_doc = fila['EMPLEADO']

        win = tk.Toplevel(self)
        win.title(f"Visor de Detalle: {tipo_doc} N° {num_doc}")
        win.geometry("1040x680")
        win.transient(self)
        win.grab_set()

        # Cabecera
        frame_head = tk.Frame(win, bg="#0056b3", padx=15, pady=10)
        frame_head.pack(fill="x", padx=10, pady=(10, 5))

        tk.Label(frame_head, text=f"📄 DETALLE DE {tipo_doc.upper()} N° {num_doc}", font=("Segoe UI", 11, "bold"), bg="#0056b3", fg="white").pack(anchor="w")
        tk.Label(frame_head, text=f"Fecha: {fecha_doc}   |   Vendedor: {vendedor_doc}   |   Cliente: {nom_cli} (RUT: {rut_cli})", font=("Segoe UI", 9), bg="#0056b3", fg="#E2E3E5").pack(anchor="w", pady=(2, 0))

        # Sección 1: Ítems de la Factura / Boleta
        frame_sec1 = tk.Frame(win, padx=10, pady=3)
        frame_sec1.pack(fill="x")
        tk.Label(frame_sec1, text=f"1. Ítems Facturados en la {tipo_doc} N° {num_doc}:", font=("Segoe UI", 9, "bold")).pack(side="left")

        btn_ajustar = tk.Button(frame_sec1, text="✏️ Ajustar % Comisión de este Documento", bg="#ffc107", fg="black", font=("Segoe UI", 8, "bold"), command=lambda: self.modal_editar_comision_seleccion())
        btn_ajustar.pack(side="right")

        cols_doc = ("PRODUCTO", "CANT", "PRECIO_FINAL", "COSTO", "TOT_VENTA", "GANANCIA", "% COMIS", "PAGO COMIS")
        tree_doc = ttk.Treeview(win, columns=cols_doc, show="headings", height=5)
        tree_doc.heading("PRODUCTO", text="Producto")
        tree_doc.heading("CANT", text="Cant")
        tree_doc.heading("PRECIO_FINAL", text="Precio ($)")
        tree_doc.heading("COSTO", text="Costo ($)")
        tree_doc.heading("TOT_VENTA", text="Total Venta ($)")
        tree_doc.heading("GANANCIA", text="Margen ($)")
        tree_doc.heading("% COMIS", text="% Comis")
        tree_doc.heading("PAGO COMIS", text="Comisión ($)")

        tree_doc.column("PRODUCTO", width=120, anchor="w")
        tree_doc.column("CANT", width=60, anchor="center")
        tree_doc.column("PRECIO_FINAL", width=110, anchor="e")
        tree_doc.column("COSTO", width=110, anchor="e")
        tree_doc.column("TOT_VENTA", width=120, anchor="e")
        tree_doc.column("GANANCIA", width=120, anchor="e")
        tree_doc.column("% COMIS", width=80, anchor="center")
        tree_doc.column("PAGO COMIS", width=120, anchor="e")

        lineas_doc = self.df_comis_detalle[self.df_comis_detalle['NUMERO'] == num_doc]
        for _, r in lineas_doc.iterrows():
            tree_doc.insert("", "end", values=(
                r['PRODUCTO'],
                f"{r['CANT']:.0f}",
                f"${r['PRECIO_FINAL']:,.0f}".replace(",", "."),
                f"${r['COSTO_UNITARIO']:,.0f}".replace(",", "."),
                f"${r['TOT_VENTA']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_TOTAL']:,.0f}".replace(",", "."),
                f"{r['PORC_COMIS']:.2f}%",
                f"${r['PAGO_EMPLEADO']:,.0f}".replace(",", ".")
            ))
        tree_doc.pack(fill="x", padx=10, pady=2)

        # Sección 2: Cruce con Notas de Crédito Asociadas al Cliente
        frame_sec2 = tk.Frame(win, padx=10, pady=(10, 2))
        frame_sec2.pack(fill="x")
        
        lbl_titulo_nc = tk.Label(frame_sec2, text="2. Notas de Crédito Asociadas a este Cliente en el Período:", font=("Segoe UI", 9, "bold"), fg="#856404")
        lbl_titulo_nc.pack(side="left")

        frame_alerta_nc = tk.Frame(win, bg="#FFF3CD", bd=1, relief="solid", padx=10, pady=4)
        lbl_alerta = tk.Label(frame_alerta_nc, text="", font=("Segoe UI", 9, "bold"), bg="#FFF3CD", fg="#856404")
        lbl_alerta.pack(anchor="w")

        cols_nc_asoc = ("TIPO", "NUM_NC", "FECHA_NC", "PRODUCTO", "CANT_DEV", "NETO_ANULADO", "MARGEN_RESTADO")
        tree_nc_asoc = ttk.Treeview(win, columns=cols_nc_asoc, show="headings", height=5)
        tree_nc_asoc.heading("TIPO", text="Tipo")
        tree_nc_asoc.heading("NUM_NC", text="N° Nota Crédito")
        tree_nc_asoc.heading("FECHA_NC", text="Fecha NC")
        tree_nc_asoc.heading("PRODUCTO", text="Producto Devuelto")
        tree_nc_asoc.heading("CANT_DEV", text="Cant Dev")
        tree_nc_asoc.heading("NETO_ANULADO", text="Venta Restada ($)")
        tree_nc_asoc.heading("MARGEN_RESTADO", text="Margen Restado ($)")

        tree_nc_asoc.column("TIPO", width=80, anchor="center")
        tree_nc_asoc.column("NUM_NC", width=105, anchor="center")
        tree_nc_asoc.column("FECHA_NC", width=90, anchor="center")
        tree_nc_asoc.column("PRODUCTO", width=120, anchor="w")
        tree_nc_asoc.column("CANT_DEV", width=70, anchor="center")
        tree_nc_asoc.column("NETO_ANULADO", width=120, anchor="e")
        tree_nc_asoc.column("MARGEN_RESTADO", width=120, anchor="e")

        try:
            conn = get_oracle_connection()
            clean_rut = str(rut_cli).upper().strip().replace('.', '').replace(' ', '').replace('/00', '').replace('-', '')
            q_nc_cli = """
            SELECT 
                'Nota Crédito' AS tipo_doc,
                d.numero_doc AS num_nc,
                TO_CHAR(d.fecha, 'YYYY-MM-DD') AS fecha_nc,
                d.cod_producto,
                ABS(d.cantidad) AS cant_dev,
                ROUND(ABS(d.cantidad) * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0)), 2) AS neto_anulado,
                ROUND((ABS(d.cantidad) * d.precio * (1 - (NVL(d.descto, 0)/100.0)) * (1 - (NVL(doc.pdesct1, 0)/100.0))) - 
                      (ABS(d.cantidad) * NVL(p.precio_costo, 0.0)), 2) AS margen_restado
            FROM detalle_venta d
            INNER JOIN documentos_venta doc 
                ON d.tipo_doc = doc.tipo_doc AND d.numero_doc = doc.numero_doc
            LEFT JOIN productos p ON d.cod_producto = p.cod_producto
            WHERE d.tipo_doc = '61'
              AND (
                  TRIM(doc.rut_cliente) = :rut_raw 
                  OR REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(doc.rut_cliente)), '.', ''), ' ', ''), '/00', ''), '-', '') = :rut_clean
              )
            ORDER BY d.fecha DESC, d.numero_doc DESC
            """
            df_nc_asoc = pd.read_sql_query(q_nc_cli, conn, params={
                'rut_raw': str(rut_cli).strip(),
                'rut_clean': clean_rut
            })
            conn.close()
            df_nc_asoc.columns = df_nc_asoc.columns.str.upper()

            if df_nc_asoc.empty:
                lbl_alerta.config(text="✅ Este cliente NO registra Notas de Crédito ni devoluciones.")
                frame_alerta_nc.pack(fill="x", padx=10, pady=3)
                tree_nc_asoc.insert("", "end", values=("INFO", "-", "-", "Sin devoluciones asociadas para este cliente", "", "", ""))
            else:
                tot_anulado = df_nc_asoc['NETO_ANULADO'].sum()
                num_ncs_unicas = len(df_nc_asoc['NUM_NC'].unique())
                lbl_alerta.config(text=f"⚠️ ATENCIÓN: Se encontraron {num_ncs_unicas} Nota(s) de Crédito para este cliente por un total restado de ${tot_anulado:,.0f}.".replace(",", "."))
                frame_alerta_nc.pack(fill="x", padx=10, pady=3)

                for _, r in df_nc_asoc.iterrows():
                    tree_nc_asoc.insert("", "end", values=(
                        r['TIPO_DOC'],
                        r['NUM_NC'],
                        r['FECHA_NC'],
                        r['COD_PRODUCTO'],
                        f"{r['CANT_DEV']:.0f}",
                        f"${r['NETO_ANULADO']:,.0f}".replace(",", "."),
                        f"${r['MARGEN_RESTADO']:,.0f}".replace(",", ".")
                    ))
        except Exception as e:
            tree_nc_asoc.insert("", "end", values=("ERROR", "-", "-", f"Error: {e}", "", "", ""))

        tree_nc_asoc.pack(fill="both", expand=True, padx=10, pady=(2, 10))

    def modal_editar_comision_seleccion(self):
        seleccionados = self.tree_det_c.selection()
        if not seleccionados:
            messagebox.showwarning("Atención", "Seleccione al menos una línea en la tabla para modificar su comisión.")
            return

        item_id = seleccionados[0]
        vals = self.tree_det_c.item(item_id, "values")
        tipo_doc, num_doc, fecha, emp, estado, cli, prod, cant, p_orig, costo, desc_p, p_fin, tot_v, tot_c, ganancia, comis_actual, pago_e, gan_emp = vals

        match = self.df_comis_detalle[(self.df_comis_detalle['NUMERO'] == num_doc) & (self.df_comis_detalle['PRODUCTO'] == prod)]
        if match.empty:
            return
        fila = match.iloc[0]

        if fila['TIPO_DOC_RAW'] == '61' or tipo_doc == 'Nota Crédito':
            messagebox.showwarning(
                "Edición No Permitida",
                f"La Nota de Crédito N° {num_doc} es un documento de anulación contable y no admite modificación manual de comisión.\n\n"
                f"Solo es posible editar porcentajes en Facturas y Boletas."
            )
            return

        if hasattr(self, 'win_edicion_comis') and self.win_edicion_comis is not None and self.win_edicion_comis.winfo_exists():
            self.win_edicion_comis.lift()
            self.win_edicion_comis.focus_force()
            return

        self.win_edicion_comis = tk.Toplevel(self)
        self.win_edicion_comis.title(f"Ajustar % Comisión - Doc: {num_doc}")
        self.win_edicion_comis.geometry("400x330")
        self.win_edicion_comis.resizable(False, False)
        self.win_edicion_comis.transient(self)
        self.win_edicion_comis.grab_set()

        tk.Label(self.win_edicion_comis, text=f"Documento: {tipo_doc} N° {num_doc}", font=("Segoe UI", 10, "bold")).pack(pady=5)
        tk.Label(self.win_edicion_comis, text=f"Vendedor: {emp}", font=("Segoe UI", 9)).pack()
        tk.Label(self.win_edicion_comis, text=f"Producto: {prod}", font=("Segoe UI", 9, "italic")).pack()
        tk.Label(self.win_edicion_comis, text=f"Ganancia / Margen de esta línea: {ganancia}", font=("Segoe UI", 9, "bold"), fg="#007bff").pack(pady=5)

        tk.Label(self.win_edicion_comis, text="Nuevo % de Comisión sobre Ganancia:").pack(pady=(5, 0))
        entry_porc = ttk.Entry(self.win_edicion_comis, width=12)
        entry_porc.insert(0, comis_actual.replace("%", ""))
        entry_porc.pack()
        entry_porc.focus_set()

        var_alcance = tk.StringVar(value="LINEA")
        rb1 = ttk.Radiobutton(self.win_edicion_comis, text=f"Aplicar solo a este producto ({prod})", variable=var_alcance, value="LINEA")
        rb1.pack(anchor="w", padx=40, pady=(10, 2))
        rb2 = ttk.Radiobutton(self.win_edicion_comis, text=f"Aplicar a TODOS los productos de la {tipo_doc} N° {num_doc}", variable=var_alcance, value="DOCUMENTO")
        rb2.pack(anchor="w", padx=40, pady=2)

        def guardar():
            try:
                nuevo_pct = float(entry_porc.get().strip())
                if var_alcance.get() == "LINEA":
                    guardar_ajuste_oracle(fila['TIPO_DOC_RAW'], num_doc, prod, fila['COD_VENDEDOR'], nuevo_pct)
                else:
                    lineas_doc = self.df_comis_detalle[self.df_comis_detalle['NUMERO'] == num_doc]
                    for _, r in lineas_doc.iterrows():
                        guardar_ajuste_oracle(r['TIPO_DOC_RAW'], num_doc, r['PRODUCTO'], r['COD_VENDEDOR'], nuevo_pct)

                self.win_edicion_comis.destroy()
                self.ejecutar_consulta_completa()
                messagebox.showinfo("Guardado", f"Ajuste del {nuevo_pct}% guardado permanentemente en Oracle.")
            except ValueError:
                messagebox.showerror("Error", "Ingrese un número válido.")

        tk.Button(self.win_edicion_comis, text="💾 Guardar en Base de Datos", bg="#28a745", fg="white", font=("Segoe UI", 9, "bold"), command=guardar).pack(pady=15)

    def crear_vista_config_especial(self):
        # 1. Panel de Comisión Especial
        frame_esp = tk.LabelFrame(self.subtab_config_c, text=" 1. Usuarios con Comisión Especial sobre Ganancia de Empresa ", font=("Segoe UI", 9, "bold"), bg="#ffffff", padx=10, pady=6)
        frame_esp.pack(fill="x", padx=15, pady=(8, 4))

        btn_add = tk.Button(frame_esp, text="➕ Asignar Usuario", bg="#28a745", fg="white", font=("Segoe UI", 8, "bold"), command=self.modal_agregar_usuario_especial)
        btn_add.pack(side="left", padx=5)

        btn_edit = tk.Button(frame_esp, text="✏️ Modificar %", bg="#ffc107", fg="black", font=("Segoe UI", 8, "bold"), command=self.modal_editar_usuario_especial)
        btn_edit.pack(side="left", padx=5)

        btn_del = tk.Button(frame_esp, text="❌ Desactivar", bg="#dc3545", fg="white", font=("Segoe UI", 8, "bold"), command=self.eliminar_usuario_especial)
        btn_del.pack(side="left", padx=5)

        cols = ("COD", "NOMBRE", "PORC_COMIS", "ESTADO")
        self.tree_cfg_esp = ttk.Treeview(frame_esp, columns=cols, show="headings", height=3)
        self.tree_cfg_esp.heading("COD", text="Cód")
        self.tree_cfg_esp.heading("NOMBRE", text="Nombre Beneficiario")
        self.tree_cfg_esp.heading("PORC_COMIS", text="% Ganancia Empresa")
        self.tree_cfg_esp.heading("ESTADO", text="Estado")

        self.tree_cfg_esp.column("COD", width=80, anchor="center")
        self.tree_cfg_esp.column("NOMBRE", width=280, anchor="w")
        self.tree_cfg_esp.column("PORC_COMIS", width=180, anchor="center")
        self.tree_cfg_esp.column("ESTADO", width=110, anchor="center")
        self.tree_cfg_esp.pack(fill="x", pady=4)

        # 2. Panel de Gestión de Vendedores
        frame_ven_pozo = tk.LabelFrame(self.subtab_config_c, text=" 2. Gestión de Vendedores (Comisión Individual vs. Pozo Común) ", font=("Segoe UI", 9, "bold"), bg="#ffffff", padx=10, pady=6)
        frame_ven_pozo.pack(fill="x", padx=15, pady=4)

        btn_toggle_pozo = tk.Button(frame_ven_pozo, text="🔄 Cambiar a Pozo / Individual", bg="#007bff", fg="white", font=("Segoe UI", 8, "bold"), command=self.toggle_vendedor_pozo)
        btn_toggle_pozo.pack(side="left", padx=5)

        btn_edit_ven = tk.Button(frame_ven_pozo, text="✏️ Modificar Tasa / Estado", bg="#ffc107", fg="black", font=("Segoe UI", 8, "bold"), command=self.modal_editar_vendedor_general)
        btn_edit_ven.pack(side="left", padx=5)

        cols_vp = ("COD", "NOMBRE", "TASA_GEN", "COMISIONA_IND")
        self.tree_cfg_ven = ttk.Treeview(frame_ven_pozo, columns=cols_vp, show="headings", height=4)
        self.tree_cfg_ven.heading("COD", text="Código")
        self.tree_cfg_ven.heading("NOMBRE", text="Nombre Vendedor")
        self.tree_cfg_ven.heading("TASA_GEN", text="Tasa General (%)")
        self.tree_cfg_ven.heading("COMISIONA_IND", text="¿Comisiona Individualmente?")

        self.tree_cfg_ven.column("COD", width=80, anchor="center")
        self.tree_cfg_ven.column("NOMBRE", width=280, anchor="w")
        self.tree_cfg_ven.column("TASA_GEN", width=140, anchor="center")
        self.tree_cfg_ven.column("COMISIONA_IND", width=220, anchor="center")

        self.tree_cfg_ven.tag_configure('ven_individual', background='#FFFFFF')
        self.tree_cfg_ven.tag_configure('ven_pozo', background='#E2E3E5', font=("Segoe UI", 9, "bold"))

        self.tree_cfg_ven.pack(fill="x", pady=4)
        self.tree_cfg_ven.bind("<Double-1>", lambda event: self.toggle_vendedor_pozo())

        # 3. Panel de Clientes TTM Service
        frame_srv_cli = tk.LabelFrame(self.subtab_config_c, text=" 3. Clientes y RUTs Asociados a TTM Service (Taller) ", font=("Segoe UI", 9, "bold"), bg="#ffffff", padx=10, pady=6)
        frame_srv_cli.pack(fill="both", expand=True, padx=15, pady=(4, 8))

        btn_add_srv = tk.Button(frame_srv_cli, text="➕ Vincular RUT a TTM Service", bg="#6f42c1", fg="white", font=("Segoe UI", 8, "bold"), command=self.modal_vincular_cliente_service)
        btn_add_srv.pack(side="left", padx=5)

        btn_del_srv = tk.Button(frame_srv_cli, text="❌ Desvincular RUT", bg="#dc3545", fg="white", font=("Segoe UI", 8, "bold"), command=self.desvincular_cliente_service)
        btn_del_srv.pack(side="left", padx=5)

        btn_rel_srv = ttk.Button(frame_srv_cli, text="🔄 Recargar Panel", command=self.cargar_tablas_configuracion)
        btn_rel_srv.pack(side="left", padx=15)

        cols_s = ("RUT", "NOMBRE", "ESTADO")
        self.tree_cfg_srv = ttk.Treeview(frame_srv_cli, columns=cols_s, show="headings", height=4)
        self.tree_cfg_srv.heading("RUT", text="RUT Cliente")
        self.tree_cfg_srv.heading("NOMBRE", text="Razón Social / Nombre")
        self.tree_cfg_srv.heading("ESTADO", text="Estado")

        self.tree_cfg_srv.column("RUT", width=140, anchor="center")
        self.tree_cfg_srv.column("NOMBRE", width=380, anchor="w")
        self.tree_cfg_srv.column("ESTADO", width=120, anchor="center")
        self.tree_cfg_srv.pack(fill="both", expand=True, pady=4)

    def toggle_vendedor_pozo(self):
        item = self.tree_cfg_ven.selection()
        if not item:
            messagebox.showwarning("Atención", "Seleccione un vendedor de la lista para cambiar su estado.")
            return
        vals = self.tree_cfg_ven.item(item[0], "values")
        cod, nom, tasa, estado = vals[0], vals[1], vals[2], vals[3]

        nuevo_estado_ind = 0 if "SÍ" in estado else 1
        nuevo_texto = "POZO COMÚN" if nuevo_estado_ind == 0 else "COMISIÓN INDIVIDUAL"

        try:
            conn = get_oracle_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE vendedores 
                SET comisiona_individual = :nuevo_estado 
                WHERE TRIM(UPPER(cod_vendedor)) = :cod
            """, {'nuevo_estado': nuevo_estado_ind, 'cod': str(cod).strip().upper()})
            conn.commit()
            cursor.close()
            conn.close()
            
            self.cargar_tablas_configuracion()
            self.recalcular_rapido_en_memoria()
            messagebox.showinfo("Actualizado", f"Estado de {nom} actualizado a {nuevo_texto}.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar en Oracle: {e}")

    def modal_editar_vendedor_general(self):
        item = self.tree_cfg_ven.selection()
        if not item:
            messagebox.showwarning("Atención", "Seleccione un vendedor.")
            return
        vals = self.tree_cfg_ven.item(item[0], "values")
        cod, nom, tasa, estado = vals[0], vals[1], vals[2], vals[3]

        win = tk.Toplevel(self)
        win.title(f"Modificar Vendedor - {nom}")
        win.geometry("380x260")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text=f"{nom} ({cod})", font=("Segoe UI", 10, "bold")).pack(pady=10)
        
        tk.Label(win, text="Tasa General de Comisión (%):").pack()
        e_tasa = ttk.Entry(win, width=12)
        e_tasa.insert(0, tasa.replace("%", "").strip())
        e_tasa.pack(pady=3)

        var_ind = tk.IntVar(value=1 if "SÍ" in estado else 0)
        chk_ind = ttk.Checkbutton(win, text="Comisiona Individualmente (Si se desmarca, va a Pozo Común)", variable=var_ind)
        chk_ind.pack(pady=12)

        def guardar():
            try:
                t_val = float(e_tasa.get().strip()) / 100.0
                ind_val = var_ind.get()
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE vendedores 
                    SET tasa_general = :tasa, comisiona_individual = :ind 
                    WHERE TRIM(UPPER(cod_vendedor)) = :cod
                """, {'tasa': t_val, 'ind': ind_val, 'cod': str(cod).strip().upper()})
                conn.commit()
                cursor.close()
                conn.close()
                win.destroy()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Actualizado", f"Configuración de {nom} actualizada en Oracle.")
            except ValueError:
                messagebox.showerror("Error", "Ingrese una tasa numérica válida.")

        tk.Button(win, text="💾 Guardar Cambios", bg="#28a745", fg="white", font=("Segoe UI", 9, "bold"), command=guardar).pack(pady=10)

    def modal_agregar_usuario_especial(self):
        try:
            conn = get_oracle_connection()
            df_v = pd.read_sql_query("SELECT cod_vendedor, nombre FROM vendedores ORDER BY nombre", conn)
            conn.close()
            df_v.columns = df_v.columns.str.upper()
            vendedores_lista = [f"{r['COD_VENDEDOR']} - {r['NOMBRE']}" for _, r in df_v.iterrows()]
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar vendedores: {e}")
            return

        win = tk.Toplevel(self)
        win.title("Asignar Comisión Especial")
        win.geometry("420x240")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Seleccione el Vendedor / Usuario:", font=("Segoe UI", 9, "bold")).pack(pady=(15, 2))
        combo_ven = ttk.Combobox(win, values=vendedores_lista, width=38, state="readonly")
        if vendedores_lista:
            combo_ven.current(0)
        combo_ven.pack(pady=5)

        tk.Label(win, text="Porcentaje sobre Ganancia Total de Empresa (%):", font=("Segoe UI", 9, "bold")).pack(pady=(10, 2))
        entry_pct = ttk.Entry(win, width=15)
        entry_pct.insert(0, "0.25")
        entry_pct.pack(pady=5)

        def guardar():
            sel = combo_ven.get()
            if not sel or " - " not in sel:
                return
            cod = sel.split(" - ")[0].strip()
            nom = sel.split(" - ")[1].strip()
            try:
                pct = float(entry_pct.get().strip())
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    MERGE INTO comision_especial_global dst
                    USING (SELECT :cod_vendedor AS cod_vendedor, :nombre AS nombre, :porc_comis AS porc_comis, 1 AS activo FROM dual) src
                    ON (dst.cod_vendedor = src.cod_vendedor)
                    WHEN MATCHED THEN
                        UPDATE SET dst.nombre = src.nombre, dst.porc_comis = src.porc_comis, dst.activo = 1
                    WHEN NOT MATCHED THEN
                        INSERT (cod_vendedor, nombre, porc_comis, activo)
                        VALUES (src.cod_vendedor, src.nombre, src.porc_comis, 1)
                """, {'cod_vendedor': cod, 'nombre': nom, 'porc_comis': pct})
                conn.commit()
                cursor.close()
                conn.close()
                win.destroy()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Éxito", f"Usuario {nom} configurado con {pct}% sobre ganancia empresa.")
            except ValueError:
                messagebox.showerror("Error", "Ingrese un número válido.")

        tk.Button(win, text="💾 Guardar en Base de Datos", bg="#28a745", fg="white", font=("Segoe UI", 9, "bold"), command=guardar).pack(pady=20)

    def modal_editar_usuario_especial(self):
        item = self.tree_cfg_esp.selection()
        if not item:
            messagebox.showwarning("Atención", "Seleccione un usuario de la lista.")
            return
        vals = self.tree_cfg_esp.item(item[0], "values")
        cod, nom, pct_str, estado = vals

        win = tk.Toplevel(self)
        win.title(f"Modificar % - {nom}")
        win.geometry("380x220")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text=f"{nom} ({cod})", font=("Segoe UI", 10, "bold")).pack(pady=15)
        tk.Label(win, text="Nuevo % sobre Ganancia Total de Empresa:").pack()
        e_pct = ttk.Entry(win, width=15)
        e_pct.insert(0, pct_str.replace("%", ""))
        e_pct.pack(pady=5)

        def guardar():
            try:
                nuevo_pct = float(e_pct.get().strip())
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE comision_especial_global 
                    SET porc_comis = :porc_comis, activo = 1 
                    WHERE cod_vendedor = :cod_vendedor
                """, {'porc_comis': nuevo_pct, 'cod_vendedor': cod})
                conn.commit()
                cursor.close()
                conn.close()
                win.destroy()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Actualizado", f"% de {nom} actualizado a {nuevo_pct}%.")
            except ValueError:
                messagebox.showerror("Error", "Ingrese un porcentaje válido.")

        tk.Button(win, text="💾 Guardar Cambios", bg="#28a745", fg="white", font=("Segoe UI", 9, "bold"), command=guardar).pack(pady=15)

    def eliminar_usuario_especial(self):
        item = self.tree_cfg_esp.selection()
        if not item:
            messagebox.showwarning("Atención", "Seleccione un usuario de la lista.")
            return
        vals = self.tree_cfg_esp.item(item[0], "values")
        cod, nom = vals[0], vals[1]

        if messagebox.askyesno("Confirmar", f"¿Desea desactivar la comisión especial para {nom}?"):
            try:
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE comision_especial_global SET activo = 0 WHERE cod_vendedor = :cod", {'cod': cod})
                conn.commit()
                cursor.close()
                conn.close()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Desactivado", f"El usuario {nom} ha sido desactivado.")
            except Exception as e:
                messagebox.showerror("Error al desactivar", f"Error: {e}")

    def exportar_excel_comisiones(self):
        if self.df_res_vendedores is None or self.df_res_vendedores.empty:
            messagebox.showwarning("Atención", "Consulte primero los datos de la temporada.")
            return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")], 
            initialfile=f"Comisiones_Vendedores_{self.entry_inicio.get()}_a_{self.entry_fin.get()}.xlsx"
        )
        if archivo:
            try:
                # Columnas ordenadas: Vendedor junto a Fecha, sin Descripción
                cols_export = [
                    'TIPO_DOC', 'NUMERO', 'FECHA_DOC', 'EMPLEADO', 'ESTADO_DOC', 'NOMBRE_CLIENTE', 'PRODUCTO',
                    'CANT', 'PRECIO_ORIG', 'COSTO_UNITARIO', 'DESC_PORC', 'PRECIO_FINAL', 
                    'TOT_VENTA', 'TOT_COSTO', 'GANANCIA_TOTAL', 
                    'PORC_COMIS', 'PAGO_EMPLEADO', 'GANANCIA_EMPRESA', 'FECHA_CALC'
                ]

                cols_trunc_res = [
                    'TOTAL_VENTAS_NETAS', 'TOTAL_NOTAS_CREDITO', 'GANANCIA_TOTAL_MARGEN', 
                    'TOTAL_COMISION_PAGAR', 'GANANCIA_NETA_EMPRESA'
                ]
                df_res_ven_exp = truncar_columnas_df(self.df_res_vendedores, cols_trunc_res)

                df_res_esp_exp = None
                if self.df_res_especial is not None and not self.df_res_especial.empty:
                    cols_trunc_esp = ['GANANCIA_BASE_EMPRESA', 'COMISION_PAGAR', 'GANANCIA_FINAL_EMPRESA']
                    df_res_esp_exp = truncar_columnas_df(self.df_res_especial, cols_trunc_esp)

                cols_trunc_det = [
                    'PRECIO_ORIG', 'COSTO_UNITARIO', 'PRECIO_FINAL', 
                    'TOT_VENTA', 'TOT_COSTO', 'GANANCIA_TOTAL', 
                    'PAGO_EMPLEADO', 'GANANCIA_EMPRESA'
                ]
                df_det_exp = truncar_columnas_df(self.df_comis_detalle[cols_export], cols_trunc_det)

                with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
                    df_res_ven_exp.to_excel(writer, sheet_name="Resumen_Vendedores", index=False)
                    if df_res_esp_exp is not None:
                        df_res_esp_exp.to_excel(writer, sheet_name="Comision_Especial_Empresa", index=False)
                    df_det_exp.to_excel(writer, sheet_name="Historial", index=False)
                    
                    ws = writer.sheets['Historial']
                    amarillo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    amarillo_font = Font(color="856404", bold=True)
                    
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols_export)):
                        tipo_doc_celda = row[0].value
                        if tipo_doc_celda == 'Nota Crédito':
                            for cell in row:
                                cell.fill = amarillo_fill
                                if cell.column == 1 or cell.column == 5:
                                    cell.font = amarillo_font

                messagebox.showinfo("Exportación Exitosa", f"Informe de comisiones guardado en:\n{archivo}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar: {e}")

    def exportar_excel_service(self):
        if self.df_service_detalle is None or self.df_service_detalle.empty:
            messagebox.showwarning("Atención", "No hay compras de TTM Service en este período.")
            return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")], 
            initialfile=f"Ventas_TTM_a_TTM_Service_{self.entry_inicio.get()}_a_{self.entry_fin.get()}.xlsx"
        )
        if archivo:
            try:
                cols_srv_exp = [
                    'TIPO_DOC', 'NUMERO', 'FECHA_DOC', 'EMPLEADO', 'ESTADO_DOC', 'NOMBRE_CLIENTE', 
                    'PRODUCTO', 'CANT', 'PRECIO_ORIG', 'COSTO_UNITARIO', 
                    'DESC_PORC', 'PRECIO_FINAL', 'TOT_VENTA', 'TOT_COSTO', 'GANANCIA_TOTAL', 'FECHA_CALC'
                ]
                
                cols_trunc_srv_res = [
                    'TOTAL_VENTAS_NETAS', 'TOTAL_NOTAS_CREDITO', 'NETO_FACTURADO', 
                    'COSTO_TOTAL_TRANSFERIDO', 'MARGEN_GANANCIA_TTM'
                ]
                df_srv_res_exp = truncar_columnas_df(self.df_service_resumen, cols_trunc_srv_res)

                cols_trunc_srv_det = [
                    'PRECIO_ORIG', 'COSTO_UNITARIO', 'PRECIO_FINAL', 
                    'TOT_VENTA', 'TOT_COSTO', 'GANANCIA_TOTAL'
                ]
                df_srv_det_exp = truncar_columnas_df(self.df_service_detalle[cols_srv_exp], cols_trunc_srv_det)

                with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
                    df_srv_res_exp.to_excel(writer, sheet_name="Resumen_Financiero_Service", index=False)
                    df_srv_det_exp.to_excel(writer, sheet_name="Detalle_Compras_Service", index=False)

                    ws = writer.sheets['Detalle_Compras_Service']
                    amarillo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    amarillo_font = Font(color="856404", bold=True)
                    
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols_srv_exp)):
                        tipo_doc_celda = row[0].value
                        if tipo_doc_celda == 'Nota Crédito':
                            for cell in row:
                                cell.fill = amarillo_fill
                                if cell.column == 1 or cell.column == 5:
                                    cell.font = amarillo_font

                messagebox.showinfo("Exportación Exitosa", f"Informe de TTM Service guardado en:\n{archivo}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar: {e}")

    def modal_vincular_cliente_service(self):
        try:
            conn = get_oracle_connection()
            df_c = pd.read_sql_query("SELECT rut_cliente, nombre FROM clientes ORDER BY nombre", conn)
            conn.close()
            df_c.columns = df_c.columns.str.upper()
            clientes_lista = [f"{r['RUT_CLIENTE']} - {r['NOMBRE']}" for _, r in df_c.iterrows()]
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar clientes: {e}")
            return

        win = tk.Toplevel(self)
        win.title("Vincular RUT a TTM Service")
        win.geometry("450x200")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Seleccione el Cliente / Razón Social de TTM Service:", font=("Segoe UI", 9, "bold")).pack(pady=(15, 5))
        combo_c = ttk.Combobox(win, values=clientes_lista, width=42)
        if clientes_lista:
            combo_c.current(0)
        combo_c.pack(pady=5)

        def guardar():
            sel = combo_c.get()
            if not sel or " - " not in sel:
                return
            rut = sel.split(" - ")[0].strip()
            nom = sel.split(" - ")[1].strip()
            try:
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    MERGE INTO ttm_service_clientes dst
                    USING (SELECT :rut_cliente AS rut_cliente, :nombre AS nombre, 1 AS activo FROM dual) src
                    ON (dst.rut_cliente = src.rut_cliente)
                    WHEN MATCHED THEN
                        UPDATE SET dst.nombre = src.nombre, dst.activo = 1
                    WHEN NOT MATCHED THEN
                        INSERT (rut_cliente, nombre, activo)
                        VALUES (src.rut_cliente, src.nombre, 1)
                """, {'rut_cliente': rut, 'nombre': nom})
                conn.commit()
                cursor.close()
                conn.close()
                win.destroy()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Éxito", f"Cliente {nom} vinculado a TTM Service.")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {e}")

        tk.Button(win, text="💾 Vincular a TTM Service", bg="#6f42c1", fg="white", font=("Segoe UI", 9, "bold"), command=guardar).pack(pady=15)

    def desvincular_cliente_service(self):
        item = self.tree_cfg_srv.selection()
        if not item:
            messagebox.showwarning("Atención", "Seleccione un cliente de la lista.")
            return
        vals = self.tree_cfg_srv.item(item[0], "values")
        rut, nom, estado = vals

        if messagebox.askyesno("Confirmar", f"¿Desea desvincular a {nom} de TTM Service?"):
            try:
                conn = get_oracle_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE ttm_service_clientes SET activo = 0 WHERE rut_cliente = :rut", {'rut': rut})
                conn.commit()
                cursor.close()
                conn.close()
                self.cargar_tablas_configuracion()
                self.recalcular_rapido_en_memoria()
                messagebox.showinfo("Desvinculado", f"{nom} ha sido desvinculado.")
            except Exception as e:
                messagebox.showerror("Error al desvincular", f"Error: {e}")

    def construir_subtabs_service(self):
        frame_top_s = tk.LabelFrame(self.tab_modulo_service, text=" Resumen Financiero Ventas a TTM Service ", font=("Segoe UI", 9, "bold"), bg="#ffffff", padx=15, pady=8)
        frame_top_s.pack(fill="x", padx=10, pady=8)

        cols_s_res = ("CONCEPTO", "VENTAS_NETAS", "NOTAS_CREDITO", "NETO_FACTURADO", "COSTO_TOTAL", "MARGEN_TTM", "ITEMS")
        self.tree_srv_res = ttk.Treeview(frame_top_s, columns=cols_s_res, show="headings", height=2)

        self.tree_srv_res.heading("CONCEPTO", text="Concepto")
        self.tree_srv_res.heading("VENTAS_NETAS", text="Ventas Netas ($)")
        self.tree_srv_res.heading("NOTAS_CREDITO", text="Notas de Crédito ($)")
        self.tree_srv_res.heading("NETO_FACTURADO", text="Total Neto Facturado ($)")
        self.tree_srv_res.heading("COSTO_TOTAL", text="Costo Total Transferido ($)")
        self.tree_srv_res.heading("MARGEN_TTM", text="Margen / Ganancia TTM ($)")
        self.tree_srv_res.heading("ITEMS", text="Total Ítems")

        self.tree_srv_res.column("CONCEPTO", width=220, anchor="w")
        self.tree_srv_res.column("VENTAS_NETAS", width=150, anchor="e")
        self.tree_srv_res.column("NOTAS_CREDITO", width=150, anchor="e")
        self.tree_srv_res.column("NETO_FACTURADO", width=170, anchor="e")
        self.tree_srv_res.column("COSTO_TOTAL", width=170, anchor="e")
        self.tree_srv_res.column("MARGEN_TTM", width=170, anchor="e")
        self.tree_srv_res.column("ITEMS", width=90, anchor="center")

        self.tree_srv_res.pack(fill="x", pady=2)

        frame_filtros_s = tk.Frame(self.tab_modulo_service, bg="#ffffff", bd=1, relief="groove", padx=10, pady=6)
        frame_filtros_s.pack(fill="x", padx=10, pady=(5, 5))

        tk.Label(frame_filtros_s, text="🔍 Buscar Factura/Boleta N°:", bg="#ffffff", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(5, 2))
        self.entry_busq_s = ttk.Entry(frame_filtros_s, width=15)
        self.entry_busq_s.pack(side="left", padx=(0, 15))
        self.entry_busq_s.bind("<KeyRelease>", lambda event: self.aplicar_filtros_service())

        btn_limp_s = ttk.Button(frame_filtros_s, text="🧹 Limpiar Filtro", command=self.limpiar_filtros_service)
        btn_limp_s.pack(side="left", padx=5)

        lbl_leyenda_s = tk.Label(frame_filtros_s, text="  Amarillo = NOTA DE CRÉDITO  ", bg="#FFF3CD", fg="#856404", font=("Segoe UI", 8, "bold"), bd=1, relief="solid")
        lbl_leyenda_s.pack(side="left", padx=15)

        self.lbl_conteo_s = tk.Label(frame_filtros_s, text="", bg="#ffffff", fg="#6c757d", font=("Segoe UI", 9, "italic"))
        self.lbl_conteo_s.pack(side="right", padx=10)

        frame_grid_s = ttk.Frame(self.tab_modulo_service)
        frame_grid_s.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        # NUEVO ORDEN EN SERVICE: Vendedor al lado de Fecha, sin Descripción
        self.cols_srv_det = [
            ("TIPO_DOC", "Tipo Doc", 90, "center"),
            ("NUMERO", "Número", 85, "center"),
            ("FECHA_DOC", "Fecha Doc", 95, "center"),
            ("EMPLEADO", "Vendedor Emisor", 160, "w"),
            ("ESTADO_DOC", "Estado Doc", 155, "center"),
            ("NOMBRE_CLIENTE", "Cliente (Service)", 180, "w"),
            ("PRODUCTO", "Producto", 110, "w"),
            ("CANT", "Cant", 55, "center"),
            ("PRECIO_ORIG", "Precio Orig ($)", 105, "e"),
            ("COSTO_UNITARIO", "Costo Unit ($)", 105, "e"),
            ("DESC_PORC", "Desc %", 70, "center"),
            ("PRECIO_FINAL", "Precio Final ($)", 110, "e"),
            ("TOT_VENTA", "Tot Venta ($)", 120, "e"),
            ("TOT_COSTO", "Tot Costo ($)", 120, "e"),
            ("GANANCIA_TOTAL", "Margen / Ganancia ($)", 130, "e")
        ]

        col_ids_s = [c[0] for c in self.cols_srv_det]
        self.tree_det_s = ttk.Treeview(frame_grid_s, columns=col_ids_s, show="headings")

        for col_id, titulo, ancho, alineacion in self.cols_srv_det:
            self.tree_det_s.heading(col_id, text=titulo)
            self.tree_det_s.column(col_id, width=ancho, anchor=alineacion)

        self.tree_det_s.tag_configure('nota_credito', background='#FFF3CD', foreground='#856404')
        self.tree_det_s.tag_configure('fila_par', background='#F8F9FA')
        self.tree_det_s.tag_configure('fila_impar', background='#FFFFFF')

        sb_y_s = ttk.Scrollbar(frame_grid_s, orient="vertical", command=self.tree_det_s.yview)
        sb_x_s = ttk.Scrollbar(frame_grid_s, orient="horizontal", command=self.tree_det_s.xview)

        self.tree_det_s.configure(yscrollcommand=sb_y_s.set, xscrollcommand=sb_x_s.set)

        self.tree_det_s.grid(row=0, column=0, sticky="nsew")
        sb_y_s.grid(row=0, column=1, sticky="ns")
        sb_x_s.grid(row=1, column=0, sticky="ew")

        frame_grid_s.grid_rowconfigure(0, weight=1)
        frame_grid_s.grid_columnconfigure(0, weight=1)

    def aplicar_filtros_comisiones(self):
        if self.df_comis_detalle is None or self.df_comis_detalle.empty:
            return

        df_f = self.df_comis_detalle.copy()

        busq = self.entry_busq_c.get().strip().lstrip('0')
        if busq:
            df_f = df_f[df_f['NUMERO'].astype(str).str.lstrip('0').str.contains(busq, case=False, na=False)]

        emp = self.combo_emp_c.get()
        if emp and emp != "TODOS":
            df_f = df_f[df_f['EMPLEADO'] == emp]

        for i in self.tree_det_c.get_children():
            self.tree_det_c.delete(i)

        for idx, (_, r) in enumerate(df_f.iterrows()):
            es_nc = (r['TIPO_DOC'] == 'Nota Crédito' or str(r['TIPO_DOC_RAW']) == '61')
            tag_fila = 'nota_credito' if es_nc else ('fila_par' if idx % 2 == 0 else 'fila_impar')

            self.tree_det_c.insert("", "end", values=(
                r['TIPO_DOC'],
                r['NUMERO'],
                r['FECHA_DOC'],
                r['EMPLEADO'],
                r['ESTADO_DOC'],
                r['NOMBRE_CLIENTE'],
                r['PRODUCTO'],
                f"{r['CANT']:.0f}",
                f"${r['PRECIO_ORIG']:,.0f}".replace(",", "."),
                f"${r['COSTO_UNITARIO']:,.0f}".replace(",", "."),
                f"{r['DESC_PORC']:.1f}%",
                f"${r['PRECIO_FINAL']:,.0f}".replace(",", "."),
                f"${r['TOT_VENTA']:,.0f}".replace(",", "."),
                f"${r['TOT_COSTO']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_TOTAL']:,.0f}".replace(",", "."),
                f"{r['PORC_COMIS']:.2f}%",
                f"${r['PAGO_EMPLEADO']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_EMPRESA']:,.0f}".replace(",", ".")
            ), tags=(tag_fila,))

        self.lbl_conteo_c.config(text=f"Mostrando {len(df_f):,} de {len(self.df_comis_detalle):,} registros")

    def limpiar_filtros_comisiones(self):
        self.entry_busq_c.delete(0, tk.END)
        self.combo_emp_c.set("TODOS")
        self.aplicar_filtros_comisiones()

    def aplicar_filtros_service(self):
        if self.df_service_detalle is None or self.df_service_detalle.empty:
            return

        df_f = self.df_service_detalle.copy()

        busq = self.entry_busq_s.get().strip().lstrip('0')
        if busq:
            df_f = df_f[df_f['NUMERO'].astype(str).str.lstrip('0').str.contains(busq, case=False, na=False)]

        for i in self.tree_det_s.get_children():
            self.tree_det_s.delete(i)

        for idx, (_, r) in enumerate(df_f.iterrows()):
            es_nc = (r['TIPO_DOC'] == 'Nota Crédito' or str(r['TIPO_DOC_RAW']) == '61')
            tag_fila = 'nota_credito' if es_nc else ('fila_par' if idx % 2 == 0 else 'fila_impar')

            self.tree_det_s.insert("", "end", values=(
                r['TIPO_DOC'],
                r['NUMERO'],
                r['FECHA_DOC'],
                r['EMPLEADO'],
                r['ESTADO_DOC'],
                r['NOMBRE_CLIENTE'],
                r['PRODUCTO'],
                f"{r['CANT']:.0f}",
                f"${r['PRECIO_ORIG']:,.0f}".replace(",", "."),
                f"${r['COSTO_UNITARIO']:,.0f}".replace(",", "."),
                f"{r['DESC_PORC']:.1f}%",
                f"${r['PRECIO_FINAL']:,.0f}".replace(",", "."),
                f"${r['TOT_VENTA']:,.0f}".replace(",", "."),
                f"${r['TOT_COSTO']:,.0f}".replace(",", "."),
                f"${r['GANANCIA_TOTAL']:,.0f}".replace(",", ".")
            ), tags=(tag_fila,))

        self.lbl_conteo_s.config(text=f"Mostrando {len(df_f):,} de {len(self.df_service_detalle):,} registros de TTM Service")

    def limpiar_filtros_service(self):
        self.entry_busq_s.delete(0, tk.END)
        self.aplicar_filtros_service()

    def ejecutar_consulta_completa(self):
        inicio_raw = self.entry_inicio.get()
        fin_raw = self.entry_fin.get()

        inicio, fin, error_msg = validar_y_normalizar_rango_fechas(inicio_raw, fin_raw)
        if error_msg:
            messagebox.showwarning("Fecha Inválida", error_msg)
            return

        self.entry_inicio.delete(0, tk.END)
        self.entry_inicio.insert(0, inicio)
        self.entry_fin.delete(0, tk.END)
        self.entry_fin.insert(0, fin)

        try:
            self.lbl_estado.config(text="⏳ Consultando Oracle Database...")
            self.update_idletasks()

            df_raw, df_v_cfg, df_e_cfg = extraer_datos_base_oracle(inicio, fin)
            self.df_raw_cache = df_raw
            self.df_ven_cfg_cache = df_v_cfg
            self.df_esp_cfg_cache = df_e_cfg

            self.recalcular_rapido_en_memoria()
            self.lbl_estado.config(text=f"🟢 Datos actualizados desde Oracle Database ({ORACLE_SERVICE}).")
        except Exception as e:
            messagebox.showerror("Error al Consultar", f"Error en Oracle: {e}")
            self.lbl_estado.config(text="Error en la consulta.")

    def recalcular_rapido_en_memoria(self):
        if self.df_raw_cache.empty:
            return

        try:
            conn = get_oracle_connection()
            self.df_ven_cfg_cache = pd.read_sql_query("SELECT cod_vendedor, nombre, tasa_general, NVL(comisiona_individual, 1) AS comisiona_individual FROM vendedores", conn)
            self.df_ven_cfg_cache.columns = self.df_ven_cfg_cache.columns.str.upper()

            self.df_esp_cfg_cache = pd.read_sql_query("SELECT * FROM comision_especial_global WHERE activo = 1", conn)
            self.df_esp_cfg_cache.columns = self.df_esp_cfg_cache.columns.str.upper()
            conn.close()
        except Exception:
            pass

        df_r_v, df_r_e, df_d_c, df_s_r, df_s_d = procesar_calculo_en_memoria(
            self.df_raw_cache, self.df_ven_cfg_cache, self.df_esp_cfg_cache
        )

        self.df_res_vendedores = df_r_v
        self.df_res_especial = df_r_e
        self.df_comis_detalle = df_d_c
        self.df_service_resumen = df_s_r
        self.df_service_detalle = df_s_d

        empleados_unicos = ["TODOS"] + sorted(list(df_d_c['EMPLEADO'].dropna().unique()))
        self.combo_emp_c['values'] = empleados_unicos

        self.repoblar_tabla_resumen()

        tot_comis_ind = df_r_v[df_r_v['COD_VENDEDOR'] != 'POZO']['TOTAL_COMISION_PAGAR'].sum() if not df_r_v.empty else 0.0
        tot_comis_pozo = df_r_v[df_r_v['COD_VENDEDOR'] == 'POZO']['TOTAL_COMISION_PAGAR'].sum() if not df_r_v.empty else 0.0
        tot_comis_especial = df_r_e['COMISION_PAGAR'].sum() if not df_r_e.empty else 0.0
        tot_comisiones_global = tot_comis_ind + tot_comis_pozo + tot_comis_especial
        
        tot_margen_periodo = df_r_v['GANANCIA_TOTAL_MARGEN'].sum() if not df_r_v.empty else 0.0
        ganancia_neta_final_empresa = tot_margen_periodo - tot_comisiones_global

        self.lbl_tot_contador.config(
            text=f"COMISIONES A PAGAR: Vendedores: ${tot_comis_ind:,.0f}  +  Pozo (2%): ${tot_comis_pozo:,.0f}  +  Especial: ${tot_comis_especial:,.0f}  =  TOTAL: ${tot_comisiones_global:,.0f}   |   MARGEN TOTAL: ${tot_margen_periodo:,.0f}   |   GANANCIA NETA EMPRESA: ${ganancia_neta_final_empresa:,.0f}".replace(",", ".")
        )

        for i in self.tree_res_e.get_children():
            self.tree_res_e.delete(i)
        if not df_r_e.empty:
            for _, r in df_r_e.iterrows():
                self.tree_res_e.insert("", "end", values=(
                    r['COD_VENDEDOR'],
                    r['BENEFICIARIO'],
                    f"${r['GANANCIA_BASE_EMPRESA']:,.0f}".replace(",", "."),
                    f"{r['PORC_COMIS_ESPECIAL']:.4f}%",
                    f"${r['COMISION_PAGAR']:,.0f}".replace(",", "."),
                    f"${r['GANANCIA_FINAL_EMPRESA']:,.0f}".replace(",", ".")
                ))

        self.aplicar_filtros_comisiones()

        for i in self.tree_srv_res.get_children():
            self.tree_srv_res.delete(i)
        if not df_s_r.empty:
            r_s = df_s_r.iloc[0]
            self.tree_srv_res.insert("", "end", values=(
                r_s['CONCEPTO'],
                f"${r_s['TOTAL_VENTAS_NETAS']:,.0f}".replace(",", "."),
                f"${r_s['TOTAL_NOTAS_CREDITO']:,.0f}".replace(",", "."),
                f"${r_s['NETO_FACTURADO']:,.0f}".replace(",", "."),
                f"${r_s['COSTO_TOTAL_TRANSFERIDO']:,.0f}".replace(",", "."),
                f"${r_s['MARGEN_GANANCIA_TTM']:,.0f}".replace(",", "."),
                f"{r_s['TOTAL_ITEMS']:,}"
            ))

        self.aplicar_filtros_service()
        self.update_idletasks()

    def verificar_conexion_inicial(self):
        try:
            conn = get_oracle_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM documentos_venta")
            total = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM clientes")
            total_cli = cursor.fetchone()[0]
            
            cursor.close()
            conn.close()
            self.lbl_estado.config(text=f"🟢 Conectado a Oracle ({ORACLE_SERVICE}) | {total:,} documentos | {total_cli:,} clientes.")
            self.cargar_tablas_configuracion()
            self.ejecutar_consulta_completa()
        except Exception as e:
            self.lbl_estado.config(text=f"🔴 Error de conexión: {e}")

    def cargar_tablas_configuracion(self):
        for item in self.tree_cfg_esp.get_children():
            self.tree_cfg_esp.delete(item)
        try:
            conn = get_oracle_connection()
            df = pd.read_sql_query("SELECT * FROM comision_especial_global ORDER BY cod_vendedor", conn)
            df.columns = df.columns.str.upper()
            for _, r in df.iterrows():
                self.tree_cfg_esp.insert("", "end", values=(
                    r['COD_VENDEDOR'],
                    r['NOMBRE'],
                    f"{r['PORC_COMIS']:.4f}%",
                    "ACTIVO" if r['ACTIVO'] == 1 else "INACTIVO"
                ))

            for item in self.tree_cfg_ven.get_children():
                self.tree_cfg_ven.delete(item)
            df_v = pd.read_sql_query("SELECT cod_vendedor, nombre, tasa_general, NVL(comisiona_individual, 1) AS comisiona_individual FROM vendedores ORDER BY nombre", conn)
            df_v.columns = df_v.columns.str.upper()
            for _, r in df_v.iterrows():
                es_ind = (r['COMISIONA_INDIVIDUAL'] == 1)
                tag_v = 'ven_individual' if es_ind else 'ven_pozo'
                self.tree_cfg_ven.insert("", "end", values=(
                    r['COD_VENDEDOR'],
                    r['NOMBRE'],
                    f"{r['TASA_GENERAL']*100:.2f}%",
                    "SÍ (Comisión Individual)" if es_ind else "NO (Va a Pozo Común)"
                ), tags=(tag_v,))

            for item in self.tree_cfg_srv.get_children():
                self.tree_cfg_srv.delete(item)
            df_s = pd.read_sql_query("SELECT * FROM ttm_service_clientes ORDER BY nombre", conn)
            conn.close()
            df_s.columns = df_s.columns.str.upper()
            for _, r in df_s.iterrows():
                self.tree_cfg_srv.insert("", "end", values=(
                    r['RUT_CLIENTE'],
                    r['NOMBRE'],
                    "ACTIVO" if r['ACTIVO'] == 1 else "INACTIVO"
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Error al leer configuración: {e}")

if __name__ == "__main__":
    app = AppComisionesOracle()
    app.mainloop()